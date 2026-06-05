"""
generar_excel_v2.py
===================
Genera MetodosNumericos_Todos_v2.xlsx replicando amburger.xlsx celda por celda.

CORRECCIONES respecto a la versión anterior:
  1. Cada hoja termina exactamente en la fila del primer SI (no bloque fijo de 25).
  2. Las iteraciones se calculan en Python primero; solo se escriben las necesarias.
  3. Steffensen reconstruido con arquitectura de 3 tablas anidadas idéntica al original.
  4. Newton 2do Orden: 2 iteraciones, no 25.
  5. f(x) escrita como =Bx^2-2 (orden correcto).
  6. Referencias $B$N para f''(x) apuntan a las celdas correctas.
  7. Newton-Raphson fila k=0: F2/G2 vacíos.
  8. Ostrowsky empieza en fila 2 (encabezados) y fila 3 (datos).

Uso:
    python3 generar_excel_v2.py
Salida:
    MetodosNumericos_Todos_v2.xlsx
"""

import math
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# ---------------------------------------------------------------------------
# CONSTANTES DEL PROBLEMA
# f(x) = x^2 - 2 = 0   →   raíz = sqrt(2) ≈ 1.41421356...
# f'(x) = 2x
# f''(x) = 2
# g(x) = (x+2)/(x+1)   (para Punto Fijo)
# ---------------------------------------------------------------------------
TOL = 0.00001   # criterio: Error% < 0.00001 → "SI"
X0 = 1.0        # punto inicial (todos los métodos)
X1 = 2.0        # segundo punto (Secante, Bisección/RF: b0)
MAX_ITER = 200  # tope de seguridad

def f(x):  return x**2 - 2
def fp(x): return 2*x
def fpp():  return 2          # constante
def g(x):  return (x + 2) / (x + 1)

def err_pct(xnew, xold):
    if xnew == 0:
        return 0.0
    return abs((xnew - xold) / xnew) * 100

# ---------------------------------------------------------------------------
# CALCULAR ITERACIONES EN PYTHON (para saber cuántas filas generar)
# Devuelve lista de dicts con los campos de cada iteración.
# ---------------------------------------------------------------------------

def calc_biseccion():
    a, b = X0, X1
    rows = [{"k": 0, "a": a, "c": (a+b)/2, "b": b,
             "fa": f(a), "fc": f((a+b)/2), "fb": f(b),
             "fafc": f(a)*f((a+b)/2), "err": None, "conv": None}]
    for k in range(1, MAX_ITER):
        c_prev = rows[-1]["c"]
        fa_prev, fc_prev = rows[-1]["fa"], rows[-1]["fc"]
        if fa_prev * fc_prev < 0:
            a, b = rows[-1]["a"], c_prev
        else:
            a, b = c_prev, rows[-1]["b"]
        c = (a + b) / 2
        fa, fc, fb = f(a), f(c), f(b)
        err = err_pct(c, c_prev)
        conv = "SI" if err < TOL else "NO"
        rows.append({"k": k, "a": a, "c": c, "b": b,
                     "fa": fa, "fc": fc, "fb": fb,
                     "fafc": fa*fc, "err": err, "conv": conv})
        if conv == "SI":
            break
    return rows

def calc_regula_falsi():
    a, b = X0, X1
    fa, fb = f(a), f(b)
    c = a - fa*(b-a)/(fb-fa)
    rows = [{"k": 0, "a": a, "c": c, "b": b,
             "fa": fa, "fc": f(c), "fb": fb,
             "fafc": fa*f(c), "err": None, "conv": None}]
    for k in range(1, MAX_ITER):
        fc_prev = rows[-1]["fc"]
        fa_prev = rows[-1]["fa"]
        if fa_prev * fc_prev < 0:
            a = rows[-1]["a"]
            b = rows[-1]["c"]
        else:
            a = rows[-1]["c"]
            b = rows[-1]["b"]
        fa, fb = f(a), f(b)
        c_prev = rows[-1]["c"]
        c = a - fa*(b-a)/(fb-fa)
        fc = f(c)
        err = err_pct(c, c_prev)
        conv = "SI" if err < TOL else "NO"
        rows.append({"k": k, "a": a, "c": c, "b": b,
                     "fa": fa, "fc": fc, "fb": fb,
                     "fafc": fa*fc, "err": err, "conv": conv})
        if conv == "SI":
            break
    return rows

def calc_punto_fijo():
    x = X0
    gx = g(x)
    rows = [{"k": 0, "x": x, "gx": gx, "err": None, "conv": None}]
    for k in range(1, MAX_ITER):
        x_prev, gx_prev = rows[-1]["x"], rows[-1]["gx"]
        x = gx_prev
        gx_new = g(x)
        err = err_pct(gx_new, x)
        conv = "SI" if err < TOL else "NO"
        rows.append({"k": k, "x": x, "gx": gx_new, "err": err, "conv": conv})
        if conv == "SI":
            break
    return rows

def calc_punto_fijo_extended(n_rows):
    """Genera n_rows iteraciones de Punto Fijo (para uso en Aitken y Steffensen)."""
    x = X0
    gx = g(x)
    rows = [{"k": 0, "x": x, "gx": gx}]
    for k in range(1, n_rows):
        x = rows[-1]["gx"]
        gx_new = g(x)
        rows.append({"k": k, "x": x, "gx": gx_new})
    return rows

def calc_aitken():
    # Aitken usa Punto Fijo como base.
    # Necesitamos las iteraciones de PF extendidas para calcular Δ².
    pf = calc_punto_fijo_extended(15)
    pk = [r["gx"] for r in pf]  # p0=g(x0), p1=g(p0), ...
    # En amburger K2=g(J2), K3=g(J3)... los xk en col J son las xk de PF
    # B(i) = K(i) - (K(i+1)-K(i))^2 / (K(i+2)-2*K(i+1)+K(i))
    # B2 usa K2,K3,K4 (indices 0,1,2 de pk)
    aitken_rows = []
    pf_rows = calc_punto_fijo_extended(13)  # need extra for denominators
    gk = [r["gx"] for r in pf_rows]  # g(xk) values: K-column
    for i in range(MAX_ITER):
        if i + 2 >= len(gk):
            break
        denom = gk[i+2] - 2*gk[i+1] + gk[i]
        if denom == 0:
            break
        phat = gk[i] - (gk[i+1] - gk[i])**2 / denom
        if i == 0:
            aitken_rows.append({"k": i, "phat": phat, "err": None, "conv": None})
        else:
            prev_phat = aitken_rows[-1]["phat"]
            err = err_pct(phat, prev_phat)
            conv = "SI" if err < TOL else "NO"
            aitken_rows.append({"k": i, "phat": phat, "err": err, "conv": conv})
            if conv == "SI":
                break
    return aitken_rows, pf_rows

def calc_newton_raphson():
    x = X0
    fx = f(x)
    fpx = fp(x)
    x1 = x - fx/fpx
    rows = [{"k": 0, "x": x, "fx": fx, "fpx": fpx, "x1": x1,
             "err": None, "conv": None}]
    for k in range(1, MAX_ITER):
        x = rows[-1]["x1"]
        fx_v = f(x)
        fpx_v = fp(x)
        x1 = x - fx_v/fpx_v
        err = err_pct(x1, x)
        conv = "SI" if err < TOL else "NO"
        rows.append({"k": k, "x": x, "fx": fx_v, "fpx": fpx_v, "x1": x1,
                     "err": err, "conv": conv})
        if conv == "SI":
            break
    return rows

def calc_newton_modificado():
    fpp_val = fpp()
    x = X0
    fx = f(x)
    fpx = fp(x)
    x1 = x - (fx*fpx)/(fpx**2 - fx*fpp_val)
    err = err_pct(x1, x)
    conv = "SI" if err < TOL else "NO"
    rows = [{"k": 0, "x": x, "fx": fx, "fpx": fpx, "fppx": fpp_val,
             "x1": x1, "err": err, "conv": conv}]
    if conv == "SI":
        return rows
    for k in range(1, MAX_ITER):
        x = rows[-1]["x1"]
        fx_v = f(x)
        fpx_v = fp(x)
        x1 = x - (fx_v*fpx_v)/(fpx_v**2 - fx_v*fpp_val)
        err = err_pct(x1, x)
        conv = "SI" if err < TOL else "NO"
        rows.append({"k": k, "x": x, "fx": fx_v, "fpx": fpx_v, "fppx": fpp_val,
                     "x1": x1, "err": err, "conv": conv})
        if conv == "SI":
            break
    return rows

def calc_newton_2do_orden():
    fpp_val = fpp()
    x = X0
    fx_v = f(x)
    fpx_v = fp(x)
    disc = fpx_v**2 - 2*fpp_val*fx_v
    x1 = x + (-fpx_v + math.sqrt(disc)) / fpp_val
    err = err_pct(x1, x)
    conv = "SI" if err < TOL else "NO"
    rows = [{"k": 0, "x": x, "fx": fx_v, "fpx": fpx_v, "fppx": fpp_val,
             "x1": x1, "err": err, "conv": conv}]
    if conv == "SI":
        return rows
    for k in range(1, MAX_ITER):
        x = rows[-1]["x1"]
        fx_v = f(x)
        fpx_v = fp(x)
        disc = fpx_v**2 - 2*fpp_val*fx_v
        x1 = x + (-fpx_v + math.sqrt(disc)) / fpp_val
        err = err_pct(x1, x)
        conv = "SI" if err < TOL else "NO"
        rows.append({"k": k, "x": x, "fx": fx_v, "fpx": fpx_v, "fppx": fpp_val,
                     "x1": x1, "err": err, "conv": conv})
        if conv == "SI":
            break
    return rows

def calc_chebyshev():
    fpp_val = fpp()
    x = X0
    for k in range(MAX_ITER):
        fx_v = f(x)
        fpx_v = fp(x)
        x1 = x - (fx_v/fpx_v) - ((fx_v**2 * fpp_val) / (2*fpx_v**3))
        err = err_pct(x1, x)
        conv = "SI" if err < TOL else "NO"
        if k == 0:
            rows = [{"k": k, "x": x, "fx": fx_v, "fpx": fpx_v, "fppx": fpp_val,
                     "x1": x1, "err": err, "conv": conv}]
        else:
            rows.append({"k": k, "x": x, "fx": fx_v, "fpx": fpx_v, "fppx": fpp_val,
                         "x1": x1, "err": err, "conv": conv})
        if conv == "SI":
            break
        x = x1
    return rows

def calc_halley():
    fpp_val = fpp()
    x = X0
    for k in range(MAX_ITER):
        fx_v = f(x)
        fpx_v = fp(x)
        x1 = x - fx_v / (fpx_v - (fpp_val*fx_v)/(2*fpx_v))
        err = err_pct(x1, x)
        conv = "SI" if err < TOL else "NO"
        if k == 0:
            rows = [{"k": k, "x": x, "fx": fx_v, "fpx": fpx_v, "fppx": fpp_val,
                     "x1": x1, "err": err, "conv": conv}]
        else:
            rows.append({"k": k, "x": x, "fx": fx_v, "fpx": fpx_v, "fppx": fpp_val,
                         "x1": x1, "err": err, "conv": conv})
        if conv == "SI":
            break
        x = x1
    return rows

def calc_super_halley():
    fpp_val = fpp()
    x = X0
    for k in range(MAX_ITER):
        fx_v = f(x)
        fpx_v = fp(x)
        numer = 2*fpx_v**2 - fx_v*fpp_val
        denom = 2*(fpx_v**2 - fx_v*fpp_val)
        x1 = x - (numer/denom)*(fx_v/fpx_v)
        err = err_pct(x1, x)
        conv = "SI" if err < TOL else "NO"
        if k == 0:
            rows = [{"k": k, "x": x, "fx": fx_v, "fpx": fpx_v, "fppx": fpp_val,
                     "x1": x1, "err": err, "conv": conv}]
        else:
            rows.append({"k": k, "x": x, "fx": fx_v, "fpx": fpx_v, "fppx": fpp_val,
                         "x1": x1, "err": err, "conv": conv})
        if conv == "SI":
            break
        x = x1
    return rows

def calc_ostrowsky():
    fpp_val = fpp()
    x = X0
    for k in range(MAX_ITER):
        fx_v = f(x)
        fpx_v = fp(x)
        denom_sq = fpx_v**2 - fx_v*fpp_val
        x1 = x - (fpx_v / math.sqrt(denom_sq)) * (fx_v / fpx_v)
        err = err_pct(x1, x)
        conv = "SI" if err < TOL else "NO"
        if k == 0:
            rows = [{"k": k, "x": x, "fx": fx_v, "fpx": fpx_v, "fppx": fpp_val,
                     "x1": x1, "err": err, "conv": conv}]
        else:
            rows.append({"k": k, "x": x, "fx": fx_v, "fpx": fpx_v, "fppx": fpp_val,
                         "x1": x1, "err": err, "conv": conv})
        if conv == "SI":
            break
        x = x1
    return rows

def calc_secante():
    x0, x1_v = X0, X1
    fx0, fx1 = f(x0), f(x1_v)
    rows = [
        {"k": 0, "x": x0, "fx": fx0, "xprev": None, "fxprev": None,
         "x1": None, "err": None, "conv": None},
        {"k": 1, "x": x1_v, "fx": fx1, "xprev": x0, "fxprev": fx0,
         "x1": x1_v - fx1*(x1_v-x0)/(fx1-fx0),
         "err": err_pct(x1_v - fx1*(x1_v-x0)/(fx1-fx0), x1_v),
         "conv": None}
    ]
    rows[1]["conv"] = "SI" if rows[1]["err"] < TOL else "NO"
    if rows[1]["conv"] == "SI":
        return rows
    for k in range(2, MAX_ITER):
        xk = rows[-1]["x1"]
        xprev = rows[-1]["x"]
        fxk = f(xk)
        fxprev = rows[-1]["fx"]
        denom = fxk - fxprev
        x_next = xk - fxk*(xk - xprev)/denom if denom != 0 else xk
        err = err_pct(x_next, xk)
        conv = "SI" if err < TOL else "NO"
        rows.append({"k": k, "x": xk, "fx": fxk, "xprev": xprev, "fxprev": fxprev,
                     "x1": x_next, "err": err, "conv": conv})
        if conv == "SI":
            break
    return rows

def calc_von_mises():
    fp0 = fp(X0)   # = 2*1 = 2, constant
    x = X0
    rows = []
    for k in range(MAX_ITER):
        fx_v = f(x)
        x1 = x - fx_v/fp0
        err = err_pct(x1, x)
        conv = "SI" if err < TOL else "NO"
        rows.append({"k": k, "x": x, "fx": fx_v, "fp0": fp0,
                     "x1": x1, "err": err, "conv": conv})
        if conv == "SI":
            break
        x = x1
    return rows

# ---------------------------------------------------------------------------
# HELPERS DE ESTILO
# ---------------------------------------------------------------------------

def make_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def thin_border():
    side = Side(border_style="thin")
    return Border(left=side, right=side, top=side, bottom=side)

def apply_header_style(cell, fill_color, font_color="1A1A2E"):
    cell.fill = make_fill(fill_color)
    cell.font = Font(name="Calibri", size=11, color=font_color)
    cell.border = thin_border()

def apply_data_style(cell, fill_color):
    cell.fill = make_fill(fill_color) if fill_color != "00000000" else PatternFill()
    cell.font = Font(name="Calibri", size=11, color="1A1A2E")
    cell.border = thin_border()

def apply_title_style(cell, fill_color, merge_range=None, ws=None):
    cell.fill = make_fill(fill_color)
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.border = thin_border()
    cell.alignment = Alignment(horizontal="center", vertical="center")

def apply_info_style(cell, fill_color="FFFFFF"):
    """Celdas de parámetros/ecuación debajo del título."""
    cell.font = Font(name="Calibri", size=11, color="1A1A2E")
    cell.border = thin_border()

# Paleta exacta de amburger
COLORS = {
    "Bisección":         {"header": "D6EAF8", "alt1": "EBF5FB", "alt2": "D6EAF8", "title": "1A5276"},
    "Regula Falsi":      {"header": "D5F5E3", "alt1": "EAFAF1", "alt2": "D5F5E3", "title": "147F57"},
    "Punto Fijo":        {"header": "E8DAEF", "alt1": "F4EBFD", "alt2": "E8DAEF", "title": "6C3483"},
    "Aitken":            {"header": "D1F2EB", "alt1": "E8F8F5", "alt2": "D1F2EB", "title": "147F57"},
    "Steffensen":        {"header": "D1F2EB", "alt1": "00000000","alt2": "D3EAFD", "title": "147F57"},
    "Newton-Raphson":    {"header": "D6DBFF", "alt1": "EEF0FF", "alt2": "D6DBFF", "title": "1F3A93"},
    "Newton Modificado": {"header": "E74C3C", "alt1": "FADBD8", "alt2": "FDEDEC", "title": "922B21"},
    "Newton 2do Orden":  {"header": "D5F5E3", "alt1": "EAFAF1", "alt2": "D5F5E3", "title": "0D553B"},
    "Chebyshev":         {"header": "D2E4F8", "alt1": "00000000","alt2": "00000000","title": "174969"},
    "Halley":            {"header": "E8F5D6", "alt1": "F0FAE8", "alt2": "E8F5D6", "title": "4D6A1F"},
    "Super Halley":      {"header": "FADBD8", "alt1": "FEF5F5", "alt2": "FADBD8", "title": "7B241C"},
    "Ostrowsky":         {"header": "E8DAEF", "alt1": "F5EEF8", "alt2": "E8DAEF", "title": "4A235A"},
    "Secante":           {"header": "D1F2EB", "alt1": "E8F8F5", "alt2": "D1F2EB", "title": "0D553B"},
    "Von Mises":         {"header": "D6EAF8", "alt1": "EBF5FB", "alt2": "D6EAF8", "title": "1A5276"},
}

def row_fill(sheet_name, data_row_idx):
    """data_row_idx: 0 = k=0 row, 1 = k=1 row, etc."""
    c = COLORS[sheet_name]
    # Even index (0,2,4...) → alt1; Odd (1,3,5...) → alt2
    return c["alt1"] if data_row_idx % 2 == 0 else c["alt2"]

def style_row(ws, excel_row, sheet_name, data_idx, n_cols):
    """Apply fill+border+font to all n_cols cells of a data row."""
    fill = row_fill(sheet_name, data_idx)
    for c in range(1, n_cols + 1):
        apply_data_style(ws.cell(excel_row, c), fill)

# ---------------------------------------------------------------------------
# WRITERS — una función por hoja
# ---------------------------------------------------------------------------

def write_biseccion(wb):
    ws = wb.create_sheet("Bisección")
    C = COLORS["Bisección"]
    rows = calc_biseccion()
    n = len(rows)

    # Row 1: headers
    headers = ["k", "aₖ", "cₖ = (a+b)/2", "bₖ", "f(aₖ)", "f(cₖ)", "f(bₖ)", "f(a)·f(c)", "Error %", "Convergencia"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        apply_header_style(cell, C["header"])

    # Row 2: k=0 (hardcoded values, no error, no convergencia)
    r = rows[0]
    ws.cell(2, 1, 0)
    ws.cell(2, 2, 1)                      # a0 hardcoded
    ws.cell(2, 3, "=(B2+D2)/2")
    ws.cell(2, 4, 2)                      # b0 hardcoded
    ws.cell(2, 5, "=B2^2-2")
    ws.cell(2, 6, "=C2^2-2")
    ws.cell(2, 7, "=D2^2-2")
    ws.cell(2, 8, "=E2*F2")
    # I2, J2 vacíos (igual que amburger)
    style_row(ws, 2, "Bisección", 0, 10)

    # Rows 3..n+1: k=1 onwards
    for k_idx, row in enumerate(rows[1:], start=1):
        er = k_idx + 2          # excel row
        prev_er = er - 1
        ws.cell(er, 1, k_idx)
        ws.cell(er, 2, f"=IF(E{prev_er}*F{prev_er}<0,B{prev_er},C{prev_er})")
        ws.cell(er, 3, f"=(B{er}+D{er})/2")
        ws.cell(er, 4, f"=IF(E{prev_er}*F{prev_er}<0,C{prev_er},D{prev_er})")
        ws.cell(er, 5, f"=B{er}^2-2")
        ws.cell(er, 6, f"=C{er}^2-2")
        ws.cell(er, 7, f"=D{er}^2-2")
        ws.cell(er, 8, f"=E{er}*F{er}")
        ws.cell(er, 9, f"=ABS((C{er}-C{prev_er})/C{er})*100")
        ws.cell(er, 10, f'=IF(I{er}<0.00001,"SI","NO")')
        style_row(ws, er, "Bisección", k_idx, 10)

    # Title row
    tr = n + 2
    ws.merge_cells(f"A{tr}:J{tr}")
    ws.cell(tr, 1, "MÉTODO DE BISECCIÓN  —  f(x) = x² − 2 = 0")
    apply_title_style(ws.cell(tr, 1), C["title"])

    # Info row
    ir = tr + 1
    ws.cell(ir, 1, "Ecuación:")
    ws.cell(ir, 2, "f(x) = x² - 2")
    ws.cell(ir, 4, "Intervalo inicial:")
    ws.cell(ir, 5, "[a₀, b₀] = [1, 2]")
    ws.cell(ir, 7, "Tolerancia:")
    ws.cell(ir, 8, "ε < 1×10⁻⁴ %")
    for ci in range(1, 11):
        apply_info_style(ws.cell(ir, ci))

    # Formula note row
    fr = ir + 1
    ws.cell(fr, 4, "FORMULA:")
    ws.cell(fr, 5, "CN = (aₙ + bₙ) / 2")


def write_regula_falsi(wb):
    ws = wb.create_sheet("Regula Falsi")
    C = COLORS["Regula Falsi"]
    rows = calc_regula_falsi()
    n = len(rows)

    headers = ["k", "aₖ", "cₖ (Reg. Falsi)", "bₖ", "f(aₖ)", "f(cₖ)", "f(bₖ)", "f(a)·f(c)", "Error %", "Convergencia"]
    for ci, h in enumerate(headers, 1):
        apply_header_style(ws.cell(1, ci, h), C["header"])

    # Row 2: k=0
    ws.cell(2, 1, 0)
    ws.cell(2, 2, 1)
    ws.cell(2, 3, "=B2-(E2*(D2-B2))/(G2-E2)")
    ws.cell(2, 4, 2)
    ws.cell(2, 5, "=B2^2-2")
    ws.cell(2, 6, "=C2^2-2")
    ws.cell(2, 7, "=D2^2-2")
    ws.cell(2, 8, "=E2*F2")
    style_row(ws, 2, "Regula Falsi", 0, 10)

    for k_idx in range(1, n):
        er = k_idx + 2
        prev = er - 1
        ws.cell(er, 1, k_idx)
        ws.cell(er, 2, f"=IF(E{prev}*F{prev}<0,B{prev},C{prev})")
        ws.cell(er, 3, f"=B{er}-(E{er}*(D{er}-B{er}))/(G{er}-E{er})")
        ws.cell(er, 4, f"=IF(E{prev}*F{prev}<0,C{prev},D{prev})")
        ws.cell(er, 5, f"=B{er}^2-2")
        ws.cell(er, 6, f"=C{er}^2-2")
        ws.cell(er, 7, f"=D{er}^2-2")
        ws.cell(er, 8, f"=E{er}*F{er}")
        ws.cell(er, 9, f"=ABS((C{er}-C{prev})/C{er})*100")
        ws.cell(er, 10, f'=IF(I{er}<0.00001,"SI","NO")')
        style_row(ws, er, "Regula Falsi", k_idx, 10)

    tr = n + 2
    ws.merge_cells(f"A{tr}:J{tr}")
    ws.cell(tr, 1, "MÉTODO DE REGULA FALSI  —  f(x) = x² − 2 = 0")
    apply_title_style(ws.cell(tr, 1), C["title"])

    ir = tr + 1
    ws.cell(ir, 1, "Ecuación:")
    ws.cell(ir, 2, "f(x) = x² - 2")
    ws.cell(ir, 4, "Intervalo inicial:")
    ws.cell(ir, 5, "[a₀, b₀] = [1, 2]")
    ws.cell(ir, 7, "Fórmula:")
    ws.cell(ir, 8, "c = a − f(a)·(b−a)/(f(b)−f(a))")
    for ci in range(1, 11):
        apply_info_style(ws.cell(ir, ci))


def write_punto_fijo(wb):
    ws = wb.create_sheet("Punto Fijo")
    C = COLORS["Punto Fijo"]
    rows = calc_punto_fijo()
    n = len(rows)

    headers = ["k", "xₖ", "g(xₖ)", "Error %", "Convergencia"]
    for ci, h in enumerate(headers, 1):
        apply_header_style(ws.cell(1, ci, h), C["header"])

    # Row 2: k=0 (no error/conv)
    ws.cell(2, 1, 0)
    ws.cell(2, 2, 1)
    ws.cell(2, 3, "=(B2+2)/(B2+1)")
    style_row(ws, 2, "Punto Fijo", 0, 5)

    for k_idx in range(1, n):
        er = k_idx + 2
        prev = er - 1
        ws.cell(er, 1, k_idx)
        ws.cell(er, 2, f"=C{prev}")
        ws.cell(er, 3, f"=(B{er}+2)/(B{er}+1)")
        ws.cell(er, 4, f"=ABS((C{er}-B{er})/C{er})*100")
        ws.cell(er, 5, f'=IF(D{er}<0.00001,"SI","NO")')
        style_row(ws, er, "Punto Fijo", k_idx, 5)

    tr = n + 2
    ws.merge_cells(f"A{tr}:F{tr}")
    ws.cell(tr, 1, "MÉTODO DE PUNTO FIJO  —  f(x) = x² − 2 = 0")
    apply_title_style(ws.cell(tr, 1), C["title"])

    ir = tr + 1
    ws.cell(ir, 1, "Ecuación:")
    ws.cell(ir, 2, "f(x) = x² - 2 → x = g(x) = (x + 2)/(x + 1)")
    ws.cell(ir, 4, "g(x):")
    ws.cell(ir, 5, "g(x) = (x + 2) / (x + 1)")
    for ci in range(1, 6):
        apply_info_style(ws.cell(ir, ci))


def write_aitken(wb):
    """
    Estructura dual exacta de amburger:
      Cols A-D: tabla Aitken (filas 2..n_aitken+1)
      Cols I-M: tabla Punto Fijo auxiliar (filas 2..n_pf+1)
      Footer Aitken en fila n_aitken+2
      Footer PF en fila (después de PF data)
    """
    ws = wb.create_sheet("Aitken")
    C = COLORS["Aitken"]

    aitken_rows, pf_rows = calc_aitken()
    na = len(aitken_rows)
    npf = len(pf_rows)

    # Row 1 headers — both tables
    for ci, h in enumerate(["k", "xk", "Error %", "Convergencia"], 1):
        apply_header_style(ws.cell(1, ci, h), C["header"])
    for ci, h in enumerate(["k", "xₖ", "g(xₖ)", "Error %", "Convergencia"], 1):
        apply_header_style(ws.cell(1, 8+ci, h), C["header"])

    # Punto Fijo sub-table (cols I=9, J=10, K=11, L=12, M=13)
    # Row 2: k=0 (no error/conv in PF)
    ws.cell(2, 9, 0)
    ws.cell(2, 10, 1)                  # J2 = x0 hardcoded
    ws.cell(2, 11, "=(J2+2)/(J2+1)")  # K2
    style_row(ws, 2, "Aitken", 0, 13)

    for k_idx in range(1, npf):
        er = k_idx + 2
        prev = er - 1
        ws.cell(er, 9, k_idx)
        ws.cell(er, 10, f"=K{prev}")
        ws.cell(er, 11, f"=(J{er}+2)/(J{er}+1)")
        ws.cell(er, 12, f"=ABS((K{er}-J{er})/K{er})*100")
        ws.cell(er, 13, f'=IF(L{er}<0.00001,"SI","NO")')
        style_row(ws, er, "Aitken", k_idx, 13)

    # Aitken main table (cols A-D)
    # B2 = K2 - (K3-K2)^2 / (K4-2*K3+K2)
    ws.cell(2, 1, 0)
    ws.cell(2, 2, "=K2-(((K3-K2)^2)/(K4-2*K3+K2))")

    for k_idx in range(1, na):
        er = k_idx + 2
        prev = er - 1
        pk = k_idx + 1  # K-row index (K2=offset 0, so Krow for aitken k=i is i+2)
        ws.cell(er, 1, k_idx)
        ws.cell(er, 2, f"=K{er}-(((K{er+1}-K{er})^2)/(K{er+2}-2*K{er+1}+K{er}))")
        ws.cell(er, 3, f"=ABS((B{er}-B{prev})/B{er})*100")
        ws.cell(er, 4, f'=IF(C{er}<0.00001,"SI","NO")')

    # Title Aitken
    tr_a = na + 2
    ws.cell(tr_a, 1, "                                                                          MÉTODO DE AITKEN (Δ²)  —  f(x) = x² − 2 = 0")
    ws.cell(tr_a, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(tr_a, 1).fill = make_fill(C["title"])

    ir_a = tr_a + 1
    ws.cell(ir_a, 1, "Ecuación:")
    ws.cell(ir_a, 2, "f(x) = x² - 2")
    ws.cell(ir_a, 4, "Método base:")
    ws.cell(ir_a, 5, "Punto Fijo  g(x) = (x+2)/(x+1)")
    ws.cell(ir_a, 6, "Nota:")
    ws.cell(ir_a, 7, "Aitken acelera la convergencia de Punto Fijo")

    ws.cell(ir_a + 1, 1, "⚠ AITKEN UTILIZA PUNTO FIJO (Método 3) como base. Se aplica la fórmula Δ²: p̂ₖ = pₖ - (pₖ₊₁-pₖ)² / (pₖ₊₂ - 2pₖ₊₁ + pₖ)")

    # PF footer (continues in col I area after data)
    tr_pf = npf + 2
    ws.cell(tr_pf, 9, "                                                                      MÉTODO DE PUNTO FIJO  —  f(x) = x² − 2 = 0")
    ws.cell(tr_pf, 9).font = Font(bold=True, color="FFFFFF")
    ws.cell(tr_pf, 9).fill = make_fill(C["title"])

    ir_pf = tr_pf + 1
    ws.cell(ir_pf, 9, "Ecuación:")
    ws.cell(ir_pf, 10, "f(x) = x² - 2 → x = g(x) = (x + 2)/(x + 1)")
    ws.cell(ir_pf, 12, "g(x):")
    ws.cell(ir_pf, 13, "g(x) = (x + 2) / (x + 1)")


def write_steffensen(wb):
    """
    Arquitectura de 3 tablas anidadas — idéntica a amburger:

      TABLA A (Steffensen principal):  cols A-D, filas 1-3
        B2 = =J8-(((J9-J8)^2)/(J10-2*J9+J8))
        B3 = =J9-(((J10-J9)^2)/(J11-2*J10+J9))

      INFO Steffensen:                 filas 4-6

      TABLA B (Aitken secundaria):     cols I-L, filas 7-11
        Headers en fila 7
        J8..J11  = Aitken sobre K20-K25
        K9..K11  = Error%

      INFO Aitken:                     filas 12-15

      INFO Punto Fijo:                 filas 17-18

      TABLA C (Punto Fijo base):       cols I-M, filas 19-29
        Headers en fila 19
        J20 = 1 (x0 hardcoded)
        K20 = =(J20+2)/(J20+1)
        J21..J29 = =K20..K28
        K21..K29 = =(J21+2)/(J21+1)
        L21..L29 = error%
        M21..M29 = convergencia
    """
    ws = wb.create_sheet("Steffensen")
    C = COLORS["Steffensen"]

    # ── TABLA A: Steffensen ──────────────────────────────────────────────
    for ci, h in enumerate(["k", "xk", "Error %", "Convergencia"], 1):
        apply_header_style(ws.cell(1, ci, h), C["header"])

    ws.cell(2, 1, 0)
    ws.cell(2, 2, "=J8-(((J9-J8)^2)/(J10-2*J9+J8))")
    # fill row 2 (k=0): no error, no conv in original
    for ci in range(1, 5):
        apply_data_style(ws.cell(2, ci), "00000000")  # no fill like amburger

    ws.cell(3, 1, 1)
    ws.cell(3, 2, "=J9-(((J10-J9)^2)/(J11-2*J10+J9))")
    ws.cell(3, 3, "=ABS((B3-B2)/B3)*100")
    ws.cell(3, 4, '=IF(C3<0.00001,"SI","NO")')
    for ci in range(1, 5):
        apply_data_style(ws.cell(3, ci), C["alt2"])

    # ── INFO Steffensen (filas 4-6) ──────────────────────────────────────
    ws.cell(4, 1, "                                                                                MÉTODO DE Steffensen   —  f(x) = x² − 2 = 0")
    ws.cell(4, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(4, 1).fill = make_fill(C["title"])
    ws.cell(4, 7, "Steffensen acelera la convergencia de Aitken")

    ws.cell(5, 1, "Ecuación:")
    ws.cell(5, 2, "f(x) = x² - 2")
    ws.cell(5, 4, "Método base:")
    ws.cell(5, 5, "Aitken")
    ws.cell(5, 6, "Nota:")
    ws.cell(6, 1, "⚠ STEFFENSEN : como base. Se aplica la fórmula Δ²: p̂ₖ = pₖ - [(pₖ₊₁-pₖ)² / (pₖ₊₂ - 2pₖ₊₁ + pₖ)]")

    # ── TABLA B: Aitken secundaria (cols I-L, filas 7-11) ────────────────
    for ci, h in enumerate(["k", "xk", "Error %", "Convergencia"], 1):
        apply_header_style(ws.cell(7, 8+ci, h), C["header"])

    # J8..J11 = Aitken Δ² aplicado sobre los K20-K23 de Punto Fijo
    ws.cell(8, 9, 0)   # I8 = k=0
    ws.cell(8, 10, "=K20-(((K21-K20)^2)/(K22-2*K21+K20))")  # J8
    for ci in range(9, 13):
        apply_data_style(ws.cell(8, ci), C["alt1"])

    ws.cell(9, 9, 1)
    ws.cell(9, 10, "=K21-(((K22-K21)^2)/(K23-2*K22+K21))")  # J9
    ws.cell(9, 11, "=ABS((J9-J8)/J9)*100")                   # K9
    ws.cell(9, 12, '=IF(K9<0.00001,"SI","NO")')              # L9
    for ci in range(9, 13):
        apply_data_style(ws.cell(9, ci), C["alt2"])

    ws.cell(10, 9, 2)
    ws.cell(10, 10, "=K22-(((K23-K22)^2)/(K24-2*K23+K22))")
    ws.cell(10, 11, "=ABS((J10-J9)/J10)*100")
    ws.cell(10, 12, '=IF(K10<0.00001,"SI","NO")')
    for ci in range(9, 13):
        apply_data_style(ws.cell(10, ci), C["alt1"])

    ws.cell(11, 9, 3)
    ws.cell(11, 10, "=K23-(((K24-K23)^2)/(K25-2*K24+K23))")
    ws.cell(11, 11, "=ABS((J11-J10)/J11)*100")
    ws.cell(11, 12, '=IF(K11<0.00001,"SI","NO")')
    for ci in range(9, 13):
        apply_data_style(ws.cell(11, ci), C["alt2"])

    # ── INFO Aitken (filas 12-15) ─────────────────────────────────────────
    ws.cell(12, 9, "                                                                                                                   MÉTODO DE AITKEN (Δ²)  —  f(x) = x² − 2 = 0")
    ws.cell(12, 9).font = Font(bold=True, color="FFFFFF")
    ws.cell(12, 9).fill = make_fill(C["title"])

    ws.cell(13, 9, "Ecuación:")
    ws.cell(13, 10, "f(x) = x² - 2")
    ws.cell(13, 12, "Método base:")
    ws.cell(13, 13, "Punto Fijo  g(x) = (x+2)/(x+1)")
    ws.cell(13, 14, "Nota:")
    ws.cell(13, 15, "Aitken acelera la convergencia de Punto Fijo")
    ws.cell(14, 9, "⚠ AITKEN UTILIZA PUNTO FIJO (Método 3) como base. Se aplica la fórmula Δ²: p̂ₖ = pₖ - [(pₖ₊₁-pₖ)² / (pₖ₊₂ - 2pₖ₊₁ + pₖ)]")
    ws.cell(15, 9, " ")

    # ── INFO Punto Fijo (filas 17-18) ─────────────────────────────────────
    ws.cell(17, 9, "                                                                                                        MÉTODO DE PUNTO FIJO  —  f(x) = x² − 2 = 0")
    ws.cell(17, 9).font = Font(bold=True, color="FFFFFF")
    ws.cell(17, 9).fill = make_fill(C["title"])

    ws.cell(18, 9, "Ecuación:")
    ws.cell(18, 10, "f(x) = x² - 2 → x = g(x) = (x + 2)/(x + 1)")
    ws.cell(18, 12, "g(x):")
    ws.cell(18, 13, "g(x) = (x + 2) / (x + 1)")

    # ── TABLA C: Punto Fijo base (cols I-M, filas 19-29) ─────────────────
    for ci, h in enumerate(["k", "xₖ", "g(xₖ)", "Error %", "Convergencia"], 1):
        apply_header_style(ws.cell(19, 8+ci, h), C["header"])

    # J20 = x0 hardcoded, K20 = g(x0)
    ws.cell(20, 9, 0)           # I20
    ws.cell(20, 10, 1)          # J20 hardcoded x0
    ws.cell(20, 11, "=(J20+2)/(J20+1)")  # K20

    for k_idx in range(1, 10):  # k=1..9 → rows 21..29
        er = 20 + k_idx
        prev = er - 1
        ws.cell(er, 9, k_idx)
        ws.cell(er, 10, f"=K{prev}")
        ws.cell(er, 11, f"=(J{er}+2)/(J{er}+1)")
        ws.cell(er, 12, f"=ABS((K{er}-J{er})/K{er})*100")
        ws.cell(er, 13, f'=IF(L{er}<0.00001,"SI","NO")')


def write_newton_raphson(wb):
    ws = wb.create_sheet("Newton-Raphson")
    C = COLORS["Newton-Raphson"]
    rows = calc_newton_raphson()
    n = len(rows)

    headers = ["k", "xₖ", "f(xₖ)", "f'(xₖ)", "xₖ₊₁", "Error %", "Convergencia"]
    for ci, h in enumerate(headers, 1):
        apply_header_style(ws.cell(1, ci, h), C["header"])

    # Row 2: k=0 — NOTE: F2 y G2 VACÍOS (igual que amburger)
    ws.cell(2, 1, 0)
    ws.cell(2, 2, 1)
    ws.cell(2, 3, "=B2^2-2")
    ws.cell(2, 4, "=2*B2")
    ws.cell(2, 5, "=B2-(C2/D2)")
    # F2, G2 vacíos — no tocar
    style_row(ws, 2, "Newton-Raphson", 0, 7)

    for k_idx in range(1, n):
        er = k_idx + 2
        prev = er - 1
        ws.cell(er, 1, k_idx)
        ws.cell(er, 2, f"=E{prev}")
        ws.cell(er, 3, f"=B{er}^2-2")
        ws.cell(er, 4, f"=2*B{er}")
        ws.cell(er, 5, f"=B{er}-(C{er}/D{er})")
        ws.cell(er, 6, f"=ABS((E{er}-B{er})/E{er})*100")
        ws.cell(er, 7, f'=IF(F{er}<0.00001,"SI","NO")')
        style_row(ws, er, "Newton-Raphson", k_idx, 7)

    tr = n + 2
    ws.merge_cells(f"A{tr}:G{tr}")
    ws.cell(tr, 1, "MÉTODO DE NEWTON-RAPHSON  —  f(x) = x² − 2 = 0")
    apply_title_style(ws.cell(tr, 1), C["title"])

    ir = tr + 1
    ws.cell(ir, 1, "Ecuación:")
    ws.cell(ir, 2, "f(x) = x² - 2,  f'(x) = 2x")
    ws.cell(ir, 4, "Fórmula:")
    ws.cell(ir, 5, "xₙ₊₁ = xₙ − f(xₙ)/f'(xₙ)")
    for ci in range(1, 8):
        apply_info_style(ws.cell(ir, ci))


def write_newton_modificado(wb):
    ws = wb.create_sheet("Newton Modificado")
    C = COLORS["Newton Modificado"]
    rows = calc_newton_modificado()
    n = len(rows)

    # f''(x) referencia: $B$9 en amburger — la celda B9 viene DESPUÉS de los datos
    # Calculamos cuántas filas de datos habrá para poner B9 en la posición correcta
    # B9 position = n + 4  (data rows n, title n+2, info n+3, fpp n+4)
    # PERO en amburger B9 = row 9 porque siempre hay exactamente 5 data rows.
    # Replicamos esa misma lógica: B{title+2} = f''
    # En amburger: n=5 datos → filas 2-6 → title=7 → info=8 → fpp=9 → $B$9
    # Fórmula: fpp_row = n + 1 + 3  = n + 4
    fpp_row = n + 1 + 3   # data ends at n+1, title at n+2, info at n+3, fpp at n+4

    headers = ["k", "xₖ", "f(xₖ)", "f'(xₖ)", "f''(xₖ)", "xₖ₊₁", "Error %", "Convergencia"]
    for ci, h in enumerate(headers, 1):
        apply_header_style(ws.cell(1, ci, h), C["header"])

    for k_idx, row in enumerate(rows):
        er = k_idx + 2
        prev = er - 1
        ws.cell(er, 1, k_idx)
        if k_idx == 0:
            ws.cell(er, 2, 1)               # x0 hardcoded
        else:
            ws.cell(er, 2, f"=F{prev}")
        ws.cell(er, 3, f"=B{er}^2-2")
        ws.cell(er, 4, f"=2*(B{er})")
        ws.cell(er, 5, f"=$B${fpp_row}")   # referencia absoluta a f''
        ws.cell(er, 6, f"=B{er}-((C{er}*D{er})/(D{er}^2-C{er}*E{er}))")
        ws.cell(er, 7, f"=ABS((F{er}-B{er})/F{er})*100")
        ws.cell(er, 8, f'=IF(G{er}<0.00001,"SI","NO")')
        style_row(ws, er, "Newton Modificado", k_idx, 8)

    tr = n + 2
    ws.merge_cells(f"A{tr}:G{tr}")
    ws.cell(tr, 1, "MÉTODO DE NEWTON MODIFICADO  —  f(x) = x² − 2 = 0")
    apply_title_style(ws.cell(tr, 1), C["title"])

    ir = tr + 1
    ws.cell(ir, 1, "Ecuación:")
    ws.cell(ir, 2, "f(x) = x² - 2,  f'(x) = 2x")
    ws.cell(ir, 4, "Fórmula:")
    ws.cell(ir, 5, "xₙ₊₁ = xₙ - [(f(xₙ)·f'(xₙ))/((f'(xₙ))² - f(xₙ)·f''(xₙ))]")
    for ci in range(1, 9):
        apply_info_style(ws.cell(ir, ci))

    # f'' row (same position as B9 in amburger)
    ws.cell(fpp_row, 1, "f''")
    ws.cell(fpp_row, 2, "=2*1")   # exactamente igual a amburger B9


def write_newton_2do_orden(wb):
    ws = wb.create_sheet("Newton 2do Orden")
    C = COLORS["Newton 2do Orden"]
    rows = calc_newton_2do_orden()
    n = len(rows)

    headers = ["k", "xₖ", "f(xₖ)", "f'(xₖ)", "f''(xₖ)", "xₖ₊₁", "Error %", "Convergencia"]
    for ci, h in enumerate(headers, 1):
        apply_header_style(ws.cell(1, ci, h), C["header"])

    for k_idx, row in enumerate(rows):
        er = k_idx + 2
        prev = er - 1
        ws.cell(er, 1, k_idx)
        if k_idx == 0:
            ws.cell(er, 2, 1)
        else:
            ws.cell(er, 2, f"=F{prev}")
        ws.cell(er, 3, f"=B{er}^2-2")
        ws.cell(er, 4, f"=2*B{er}")
        ws.cell(er, 5, 2)    # f''(x)=2 hardcoded como valor literal
        ws.cell(er, 6, f"=B{er}+((-D{er}+((D{er})^2-2*E{er}*C{er})^(1/2))/E{er})")
        ws.cell(er, 7, f"=ABS((F{er}-B{er})/F{er})*100")
        ws.cell(er, 8, f'=IF(G{er}<0.00001,"SI","NO")')
        style_row(ws, er, "Newton 2do Orden", k_idx, 8)

    tr = n + 2
    ws.merge_cells(f"A{tr}:H{tr}")
    ws.cell(tr, 1, "MÉTODO DE NEWTON 2do ORDEN  —  f(x) = x² − 2 = 0")
    apply_title_style(ws.cell(tr, 1), C["title"])

    ir = tr + 1
    ws.cell(ir, 1, "Ecuación:")
    ws.cell(ir, 2, "f(x) = x² - 2,  f'(x) = 2x,  f''(x) = 2")
    ws.cell(ir, 5, "Fórmula:")
    ws.cell(ir, 6, "x_{n+1} = x_n + (-f'(x_n) + sqrt((f'(x_n))^2 - 2*f''(x_n)*f(x_n))) / f''(x_n)")
    for ci in range(1, 9):
        apply_info_style(ws.cell(ir, ci))


def _write_derivative2_method(wb, sheet_name, method_title, formula_text,
                               calc_func, fpp_col_ref, fpp_row_offset,
                               xnext_formula_fn):
    """
    Escribe un método con f''(x) constante (Chebyshev, Halley, Super Halley).
    fpp_row_offset: cuántas filas después del último dato está la celda de f'' (en amburger).
      Chebyshev B8   → offset 3 desde title (title=n+2, info=n+3, fpp=n+4) ... wait
      Chebyshev: 4 data rows → 2-5, title=6, info=7, fpp=8 → offset desde title = 2
      Halley:    4 data rows → same → B8 → offset 2
      Super Halley: 3 data rows → 2-4, title=5, info=6, fpp=7 → B7 → offset 2
    All three: fpp is at row n+4 (n data rows, n+2 title, n+3 info, n+4 fpp)
    Actually checking amburger:
      Chebyshev: n=4, last data row=5, title=6, info=7, fpp=8  → n+4
      Halley:    n=4, last data row=5, title=6, info=7, fpp=8  → n+4
      Super Halley: n=3, last data row=4, title=5, info=6, fpp=7 → n+4
    So fpp_row = n + 4 always for these three.
    """
    ws = wb.create_sheet(sheet_name)
    C = COLORS[sheet_name]
    rows = calc_func()
    n = len(rows)
    fpp_row = n + 4

    headers = ["k", "xₖ", "f(xₖ)", "f'(xₖ)", "f''(xₖ)", "xₖ₊₁", "Error %", "Convergencia"]
    for ci, h in enumerate(headers, 1):
        apply_header_style(ws.cell(1, ci, h), C["header"])

    for k_idx, row in enumerate(rows):
        er = k_idx + 2
        prev = er - 1
        ws.cell(er, 1, k_idx)
        if k_idx == 0:
            ws.cell(er, 2, 1)
        else:
            ws.cell(er, 2, f"=F{prev}")
        ws.cell(er, 3, f"=B{er}^2-2")
        ws.cell(er, 4, f"=2*B{er}") if sheet_name != "Newton Modificado" else ws.cell(er, 4, f"=2*(B{er})")
        ws.cell(er, 5, f"=$B${fpp_row}")
        ws.cell(er, 6, xnext_formula_fn(er))
        ws.cell(er, 7, f"=ABS((F{er}-B{er})/F{er})*100")
        ws.cell(er, 8, f'=IF(G{er}<0.00001,"SI","NO")')
        style_row(ws, er, sheet_name, k_idx, 8)

    tr = n + 2
    ws.merge_cells(f"A{tr}:H{tr}")
    ws.cell(tr, 1, method_title)
    apply_title_style(ws.cell(tr, 1), C["title"])

    ir = tr + 1
    ws.cell(ir, 1, "Ecuación:")
    ws.cell(ir, 2, "f(x) = x² - 2,  f'(x) = 2x,  f''(x) = 2")
    ws.cell(ir, 5, "Fórmula:")
    ws.cell(ir, 6, formula_text)
    for ci in range(1, 9):
        apply_info_style(ws.cell(ir, ci))

    # f'' hardcoded cell
    ws.cell(fpp_row, 1, "f''(x)")
    ws.cell(fpp_row, 2, 2)


def write_chebyshev(wb):
    _write_derivative2_method(
        wb, "Chebyshev",
        "MÉTODO DE CHEBYSHEV  —  f(x) = x² − 2 = 0",
        "xₙ₊₁ = xₙ − (f(xₙ) / f′(xₙ)) − [ (f(xₙ))² · f″(xₙ) ] / [ 2 · (f′(xₙ))³ ]",
        calc_chebyshev, "B", 0,
        lambda er: f"=B{er}-(C{er}/D{er})-(((C{er}^2)*(E{er}))/(2*(D{er})^3))"
    )


def write_halley(wb):
    _write_derivative2_method(
        wb, "Halley",
        "MÉTODO DE HALLEY  —  f(x) = x² − 2 = 0",
        "xₙ₊₁ = xₙ − f / [f' − (f''·f)/(2·f')]",
        calc_halley, "B", 0,
        lambda er: f"=B{er}-(C{er}/(D{er}-(E{er}*C{er})/(2*D{er})))"
    )


def write_super_halley(wb):
    _write_derivative2_method(
        wb, "Super Halley",
        "MÉTODO DE SUPER HALLEY  —  f(x) = x² − 2 = 0",
        "xₙ₊₁ = xₙ − [ 2·(f′(xₙ))² − f(xₙ)·f″(xₙ) ] / [ 2·((f′(xₙ))² − f(xₙ)·f″(xₙ)) ] · (f(xₙ) / f′(xₙ))",
        calc_super_halley, "B", 0,
        lambda er: f"=B{er}-(2*(D{er}^2)-C{er}*E{er})/(2*((D{er}^2)-C{er}*E{er}))*((C{er})/(D{er}))"
    )


def write_ostrowsky(wb):
    """
    Ostrowsky empieza en fila 2 (encabezados), datos desde fila 3.
    E es f''(x) = 2 hardcoded como valor en cada fila (no referencia).
    """
    ws = wb.create_sheet("Ostrowsky")
    C = COLORS["Ostrowsky"]
    rows = calc_ostrowsky()
    n = len(rows)

    # Fila 1 vacía (igual que amburger)
    # Fila 2: headers
    headers = ["k", "xₖ", "f(xₖ)", "f'(xₖ)", "f''(xₖ)", "xₖ₊₁", "Error %", "Convergencia"]
    for ci, h in enumerate(headers, 1):
        apply_header_style(ws.cell(2, ci, h), C["header"])

    # Datos desde fila 3
    for k_idx, row in enumerate(rows):
        er = k_idx + 3
        prev = er - 1
        ws.cell(er, 1, k_idx)
        if k_idx == 0:
            ws.cell(er, 2, 1)   # x0 hardcoded
        else:
            ws.cell(er, 2, f"=F{prev}")
        ws.cell(er, 3, f"=B{er}^2-2")
        ws.cell(er, 4, f"=2*B{er}")
        ws.cell(er, 5, 2)      # f''(x) = 2 hardcoded como valor
        ws.cell(er, 6, f"=B{er}-(D{er}/((D{er}^2-C{er}*E{er})^(1/2)))*(C{er}/D{er})")
        ws.cell(er, 7, f"=ABS((F{er}-B{er})/F{er})*100")
        ws.cell(er, 8, f'=IF(G{er}<0.00001,"SI","NO")')
        style_row(ws, er, "Ostrowsky", k_idx, 8)

    tr = n + 3
    ws.merge_cells(f"A{tr}:H{tr}")
    ws.cell(tr, 1, "MÉTODO DE OSTROWSKY  —  f(x) = x² − 2 = 0")
    apply_title_style(ws.cell(tr, 1), C["title"])

    ir = tr + 1
    ws.cell(ir, 1, "Ecuación:")
    ws.cell(ir, 2, "f(x) = x² - 2,  f'(x) = 2x,  f''(x) = 2")
    ws.cell(ir, 5, "Fórmula:")
    ws.cell(ir, 6, "xₙ₊₁ = xₙ − (f'(xₙ)/√(f'²−f·f'')) · (f/f')")
    for ci in range(1, 9):
        apply_info_style(ws.cell(ir, ci))


def write_secante(wb):
    ws = wb.create_sheet("Secante")
    C = COLORS["Secante"]
    rows = calc_secante()
    n = len(rows)

    headers = ["k", "xₖ", "f(xₖ)", "xₖ₋₁ (prev)", "f(xₖ-₁)", "xₖ₊₁", "Error %", "Convergencia"]
    for ci, h in enumerate(headers, 1):
        apply_header_style(ws.cell(1, ci, h), C["header"])

    # Row 2: k=0 (solo B2 y C2)
    ws.cell(2, 1, 0)
    ws.cell(2, 2, 1)
    ws.cell(2, 3, "=B2^2-2")
    style_row(ws, 2, "Secante", 0, 8)

    # Row 3: k=1 (B3=2 hardcoded, D3=B2, E3=C2)
    ws.cell(3, 1, 1)
    ws.cell(3, 2, 2)
    ws.cell(3, 3, "=B3^2-2")
    ws.cell(3, 4, "=B2")
    ws.cell(3, 5, "=C2")
    ws.cell(3, 6, "=B3-(C3*(B3-D3)/(C3-E3))")
    ws.cell(3, 7, "=IFERROR(ABS((F3-B3)/F3)*100,0)")
    ws.cell(3, 8, '=IF(IFERROR(G3,0)<0.00001,"SI","NO")')
    style_row(ws, 3, "Secante", 1, 8)

    # Rows 4..n+1: k=2 onwards with IFERROR
    for k_idx in range(2, n):
        er = k_idx + 2
        prev = er - 1
        ws.cell(er, 1, k_idx)
        ws.cell(er, 2, f"=IFERROR(F{prev},B{prev})")
        ws.cell(er, 3, f"=IFERROR(B{er}^2-2,0)")
        ws.cell(er, 4, f"=IFERROR(B{prev},B{prev})")
        ws.cell(er, 5, f"=IFERROR(C{prev},0)")
        ws.cell(er, 6, f"=IFERROR(B{er}-(C{er}*(B{er}-D{er})/(C{er}-E{er})),B{er})")
        ws.cell(er, 7, f"=IFERROR(ABS((F{er}-B{er})/F{er})*100,0)")
        ws.cell(er, 8, f'=IF(IFERROR(G{er},0)<0.00001,"SI","NO")')
        style_row(ws, er, "Secante", k_idx, 8)

    tr = n + 2
    ws.merge_cells(f"A{tr}:H{tr}")
    ws.cell(tr, 1, "MÉTODO DE LA SECANTE  —  f(x) = x² − 2 = 0")
    apply_title_style(ws.cell(tr, 1), C["title"])

    ir = tr + 1
    ws.cell(ir, 1, "Ecuación:")
    ws.cell(ir, 2, "f(x) = x² - 2")
    ws.cell(ir, 4, "Puntos iniciales:")
    ws.cell(ir, 5, "x₀ = 1,  x₁ = 2")
    ws.cell(ir, 6, "Fórmula:")
    ws.cell(ir, 7, "xₙ₊₁ = xₙ − f(xₙ)·(xₙ−xₙ₋₁)/(f(xₙ)−f(xₙ₋₁))")
    for ci in range(1, 9):
        apply_info_style(ws.cell(ir, ci))


def write_von_mises(wb):
    ws = wb.create_sheet("Von Mises")
    C = COLORS["Von Mises"]
    rows = calc_von_mises()
    n = len(rows)

    headers = ["k", "xₖ", "f(xₖ)", "f'(x0)", "xₖ₊₁", "Error %", "Convergencia"]
    for ci, h in enumerate(headers, 1):
        apply_header_style(ws.cell(1, ci, h), C["header"])

    # D column = f'(x0) = 2 constante en todas las filas
    for k_idx, row in enumerate(rows):
        er = k_idx + 2
        prev = er - 1
        ws.cell(er, 1, k_idx)
        if k_idx == 0:
            ws.cell(er, 2, 1)
        else:
            ws.cell(er, 2, f"=E{prev}")
        ws.cell(er, 3, f"=B{er}^2-2")
        ws.cell(er, 4, 2)        # f'(x0) constante
        ws.cell(er, 5, f"=B{er}-(C{er}/D{er})")
        ws.cell(er, 6, f"=ABS((E{er}-B{er})/E{er})*100")
        ws.cell(er, 7, f'=IF(F{er}<0.00001,"SI","NO")')
        style_row(ws, er, "Von Mises", k_idx, 7)

    tr = n + 2
    ws.merge_cells(f"A{tr}:H{tr}")
    ws.cell(tr, 1, "MÉTODO DE VON MISES  —  f(x) = x² − 2 = 0")
    apply_title_style(ws.cell(tr, 1), C["title"])

    ir = tr + 1
    ws.cell(ir, 1, "Ecuación:")
    ws.cell(ir, 2, "f(x) = x² - 2,  f'(x) = 2x")
    ws.cell(ir, 4, "f'(x₀):")
    ws.cell(ir, 5, "f'(1) = 2  (constante en todas las iteraciones)")
    ws.cell(ir, 6, "Fórmula:")
    ws.cell(ir, 7, "xₙ₊₁ = xₙ − f(xₙ) / f'(x₀)")
    for ci in range(1, 8):
        apply_info_style(ws.cell(ir, ci))


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Generando hojas...")
    write_biseccion(wb);       print("  ✓ Bisección")
    write_regula_falsi(wb);    print("  ✓ Regula Falsi")
    write_punto_fijo(wb);      print("  ✓ Punto Fijo")
    write_aitken(wb);          print("  ✓ Aitken")
    write_steffensen(wb);      print("  ✓ Steffensen")
    write_newton_raphson(wb);  print("  ✓ Newton-Raphson")
    write_newton_modificado(wb); print("  ✓ Newton Modificado")
    write_newton_2do_orden(wb);  print("  ✓ Newton 2do Orden")
    write_chebyshev(wb);       print("  ✓ Chebyshev")
    write_halley(wb);          print("  ✓ Halley")
    write_super_halley(wb);    print("  ✓ Super Halley")
    write_ostrowsky(wb);       print("  ✓ Ostrowsky")
    write_secante(wb);         print("  ✓ Secante")
    write_von_mises(wb);       print("  ✓ Von Mises")

    out = "/mnt/user-data/outputs/MetodosNumericos_Todos_v2.xlsx"
    wb.save(out)
    print(f"\nGuardado: {out}")

    # Quick verification
    verify(out)


def verify(path):
    """Verificación rápida de iteraciones por hoja."""
    import openpyxl as xl
    wb = xl.load_workbook(path, data_only=False)
    print("\n=== VERIFICACIÓN DE FILAS POR MÉTODO ===")
    expected = {
        "Bisección": 23, "Regula Falsi": 10, "Punto Fijo": 10,
        "Aitken": 4, "Steffensen": 2, "Newton-Raphson": 5,
        "Newton Modificado": 5, "Newton 2do Orden": 2, "Chebyshev": 4,
        "Halley": 4, "Super Halley": 3, "Ostrowsky": 3,
        "Secante": 7, "Von Mises": 18
    }
    all_ok = True
    for sh in wb.sheetnames:
        ws = wb[sh]
        data_rows = []
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                data_rows.append(r)
        n = len(data_rows)
        exp = expected.get(sh, "?")
        ok = "✓" if n == exp else "✗"
        if n != exp:
            all_ok = False
        print(f"  {ok} {sh}: {n} iteraciones (esperado: {exp})")
    if all_ok:
        print("\n✓ TODAS LAS HOJAS TIENEN EL NÚMERO CORRECTO DE ITERACIONES")
    else:
        print("\n✗ ALGUNAS HOJAS TIENEN DIFERENCIAS")


if __name__ == "__main__":
    main()