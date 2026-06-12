"""
excel_generator.py
------------------
Public API for generating Excel workbooks with real formula-based sheets.

Entry points:
  generate_single(method_key, ...)  → bytes (xlsx)
  generate_all(...)                 → bytes (xlsx, 14 sheets)

Both functions return the raw .xlsx bytes, ready to be written to disk
or streamed via FastAPI.

Integration contract with existing backend modules:
  - Expects fx_str as a string parseable by SymPy (e.g. "x**2 - 2")
  - Derivatives are computed internally via sympy_to_excel.py
  - params dict keys must match what excel_templates.py expects
  - Does NOT call biseccion.py / newton_raphson.py etc. — it generates
    Excel files independently using the same mathematical definitions.

Sheet order matches amburger.xlsx:
  1  Bisección           8  Newton 2do Orden
  2  Regula Falsi        9  Chebyshev
  3  Punto Fijo         10  Halley
  4  Aitken             11  Super Halley
  5  Steffensen         12  Ostrowsky
  6  Newton-Raphson     13  Secante
  7  Newton Modificado  14  Von Mises
"""

from __future__ import annotations
import io
from typing import Any

from openpyxl import Workbook

from backend.core.sympy_to_excel import expr_to_sympy, to_excel_formula, derivative_to_excel
from backend.excel.excel_templates import TEMPLATE_REGISTRY, build_status_panel

# Runners de los métodos numéricos — usados SOLO para calcular cuántas filas de
# iteración generar (para que cada tabla termine en la fila del primer "SI").
# Las fórmulas de Excel siguen siendo la fuente de verdad de los valores.
from backend.methods.newton_raphson import run as _run_newton
from backend.methods.biseccion import run as _run_biseccion
from backend.methods.regula_falsi import run as _run_regula_falsi
from backend.methods.secante import run as _run_secante
from backend.methods.punto_fijo import run as _run_punto_fijo
from backend.methods.aitken import run as _run_aitken
from backend.methods.steffensen import run as _run_steffensen
from backend.methods.von_mises import run as _run_von_mises
from backend.methods.newton_family import (
    run_newton_modificado as _run_newton_modificado,
    run_newton_segundo_orden as _run_newton_2do,
    run_chebyshev as _run_chebyshev,
    run_halley as _run_halley,
    run_super_halley as _run_super_halley,
    run_ostrowsky as _run_ostrowsky,
)

# method_key (igual que TEMPLATE_REGISTRY) -> runner(eq, params) -> MethodResult
_METHOD_RUNNERS = {
    "newton_raphson":    lambda eq, p: _run_newton(eq, p),
    "biseccion":         lambda eq, p: _run_biseccion(eq, p),
    "regula_falsi":      lambda eq, p: _run_regula_falsi(eq, p),
    "secante":           lambda eq, p: _run_secante(eq, p),
    "punto_fijo":        lambda eq, p: _run_punto_fijo(eq, p),
    "aitken":            lambda eq, p: _run_aitken(eq, p),
    "steffensen":        lambda eq, p: _run_steffensen(eq, p),
    "von_mises":         lambda eq, p: _run_von_mises(eq, p),
    "newton_modificado": lambda eq, p: _run_newton_modificado(eq, p),
    "newton_2do_orden":  lambda eq, p: _run_newton_2do(eq, p),
    "chebyshev":         lambda eq, p: _run_chebyshev(eq, p),
    "halley":            lambda eq, p: _run_halley(eq, p),
    "super_halley":      lambda eq, p: _run_super_halley(eq, p),
    "ostrowsky":         lambda eq, p: _run_ostrowsky(eq, p),
}


def _method_status(method_key: str, fx_str: str):
    """
    Ejecuta el método numérico ya existente (mismos seeds que el Excel vía
    auto_params) y devuelve (applicable, converged, reason, iteration_count),
    o None si el método no tiene runner o falla al evaluarse.
    """
    runner = _METHOD_RUNNERS.get(method_key)
    if runner is None:
        return None
    try:
        from backend.core.equation_parser import parse_equation
        from backend.core.auto_params import generate_params
        eq = parse_equation(fx_str)
        params = generate_params(eq)
        res = runner(eq, params)
    except Exception:
        return None
    return (
        bool(getattr(res, "applicable", True)),
        bool(getattr(res, "converged", False)),
        str(getattr(res, "reason", "") or ""),
        getattr(res, "iteration_count", None),
    )


def _has_real_roots(fx_str: str):
    """
    True / False / None(desconocido): ¿la ecuación tiene alguna raíz real?
    Solo afirma False cuando SymPy resuelve y TODAS las soluciones son complejas
    (caso claro como x^2+1). Si no puede resolver simbólicamente, devuelve None.
    """
    try:
        import sympy as sp
        x = sp.Symbol("x")
        expr = sp.sympify(fx_str.replace("^", "**"), locals={"x": x})
        sols = sp.solve(expr, x)
    except Exception:
        return None
    if not sols:
        return None
    reals = [s for s in sols if getattr(s, "is_real", None) is True]
    return len(reals) > 0


def _sized_n_iter(method_key: str, iteration_count, default_n_iter: int) -> int:
    """
    Filas de iteración para que la tabla TERMINE en la fila del "SI". Verificado:
    iteration_count coincide EXACTAMENTE con la fila del "SI" del Excel.
      - Secante: k=0 y k=1 fuera del loop; loop llega a k=n_iter -> n_iter = itc.
      - El resto: genera k=0..n_iter-1 -> n_iter = itc + 1.
    """
    if iteration_count is None:
        return default_n_iter
    if method_key == "secante":
        return max(int(iteration_count), 1)
    return int(iteration_count) + 1


# ---------------------------------------------------------------------------
# Sheet display names (preserves exact naming from amburger.xlsx)
# ---------------------------------------------------------------------------

SHEET_NAMES: dict[str, str] = {
    "biseccion":         "Bisección",
    "regula_falsi":      "Regula Falsi",
    "punto_fijo":        "Punto Fijo",
    "aitken":            "Aitken",
    "steffensen":        "Steffensen",
    "newton_raphson":    "Newton-Raphson",
    "newton_modificado": "Newton Modificado",
    "newton_2do_orden":  "Newton 2do Orden",
    "chebyshev":         "Chebyshev",
    "halley":            "Halley",
    "super_halley":      "Super Halley",
    "ostrowsky":         "Ostrowsky",
    "secante":           "Secante",
    "von_mises":         "Von Mises",
}

# Ordered list for full-workbook export
ALL_METHODS: list[str] = [
    "biseccion", "regula_falsi", "punto_fijo",
    "aitken", "steffensen",
    "newton_raphson", "newton_modificado", "newton_2do_orden",
    "chebyshev", "halley", "super_halley", "ostrowsky",
    "secante", "von_mises",
]


# ---------------------------------------------------------------------------
# Derivative helpers
# ---------------------------------------------------------------------------

def _compute_derivatives(fx_str: str) -> tuple[str, str]:
    """Return (fpx_expr_str, fppx_expr_str) as SymPy-parseable strings."""
    import sympy as sp
    from sympy import Symbol
    xsym = Symbol("x")
    expr = expr_to_sympy(fx_str)
    fp  = sp.diff(expr, xsym, 1)
    fpp = sp.diff(expr, xsym, 2)
    return str(fp), str(fpp)


# ---------------------------------------------------------------------------
# Default parameter resolver
# (integrates with auto_params.py when available)
# ---------------------------------------------------------------------------

def _default_params(method_key: str, fx_str: str) -> dict[str, Any]:
    """
    Generate sensible default parameters using auto_params.
    """
    from backend.core.equation_parser import parse_equation
    from backend.core.auto_params import generate_params

    eq = parse_equation(fx_str)
    p  = generate_params(eq)

    interval_methods  = {"biseccion", "regula_falsi"}
    two_point_methods = {"secante"}
    g_methods         = {"punto_fijo", "aitken", "steffensen"}

    gx_str     = str(p.gx_sympy) if p.gx_sympy is not None else "x"
    gx_display = f"g(x) = {gx_str}"

    if method_key in interval_methods:
        return {"a0": p.a, "b0": p.b}
    if method_key in two_point_methods:
        return {"x0": p.x0, "x1": p.x0_alt}
    if method_key in g_methods:
        return {"x0": p.x0, "g_str": gx_str, "g_display": gx_display}
    return {"x0": p.x0}


# ---------------------------------------------------------------------------
# Core sheet builder
# ---------------------------------------------------------------------------

def _build_sheet(wb: Workbook, method_key: str, fx_str: str,
                 params: dict | None, n_iter: int, eq_label: str) -> None:
    """Add a single method sheet to wb."""
    if params is None:
        params = _default_params(method_key, fx_str)

    sheet_name = SHEET_NAMES[method_key]
    ws = wb.create_sheet(title=sheet_name)

    # Si el método NO es aplicable o NO converge para esta ecuación, mostrar un
    # panel explicativo (con el motivo que ya calcula el backend) en lugar de una
    # tabla de fórmulas que se llenaría de #NUM!/#DIV/0!.
    status = _method_status(method_key, fx_str)
    if status is not None:
        applicable, converged, reason, itc = status
        if (not applicable) or (not converged):
            build_status_panel(ws, method_key, sheet_name, eq_label or fx_str,
                                applicable, converged, reason, fx_str,
                                _has_real_roots(fx_str))
            return
        # Convergió: dimensionar la tabla para que termine en la fila del "SI".
        n_iter = _sized_n_iter(method_key, itc, n_iter)

    fpx_str, fppx_str = _compute_derivatives(fx_str)

    template_cls = TEMPLATE_REGISTRY[method_key]
    template = template_cls()

    fp_methods = {"newton_raphson", "von_mises"}
    fpp_methods = {
        "newton_modificado", "newton_2do_orden",
        "chebyshev", "halley", "super_halley", "ostrowsky",
    }

    if method_key in fpp_methods:
        template.build(
            ws=ws, fx_str=fx_str, params=params,
            n_iter=n_iter, eq_label=eq_label,
            fpx_str=fpx_str, fppx_str=fppx_str,
        )
    elif method_key in fp_methods:
        template.build(
            ws=ws, fx_str=fx_str, params=params,
            n_iter=n_iter, eq_label=eq_label,
            fpx_str=fpx_str,
        )
    else:
        template.build(
            ws=ws, fx_str=fx_str, params=params,
            n_iter=n_iter, eq_label=eq_label,
        )


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_single(
    method_key: str,
    fx_str: str,
    *,
    params: dict | None = None,
    n_iter: int = 25,
    eq_label: str = "",
) -> bytes:
    """
    Generate a single-sheet xlsx for the given method.

    Parameters
    ----------
    method_key : str
        One of the keys in SHEET_NAMES.
    fx_str : str
        f(x) as SymPy-parseable string, e.g. "x**2 - 2"
    params : dict, optional
        Method-specific initial values. Auto-generated if not provided.
    n_iter : int
        Number of iteration rows (default 25).
    eq_label : str
        Human-readable equation for footer, e.g. "f(x) = x² - 2"

    Returns
    -------
    bytes
        Raw .xlsx content ready for download.
    """
    if method_key not in TEMPLATE_REGISTRY:
        raise ValueError(
            f"Unknown method: '{method_key}'. "
            f"Valid options: {list(TEMPLATE_REGISTRY.keys())}"
        )

    wb = Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    _build_sheet(wb, method_key, fx_str, params, n_iter, eq_label or fx_str)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_all(
    fx_str: str,
    *,
    params_per_method: dict[str, dict] | None = None,
    n_iter: int = 25,
    eq_label: str = "",
    methods: list[str] | None = None,
) -> bytes:
    """
    Generate a multi-sheet xlsx with all 14 methods (or a subset).

    Parameters
    ----------
    fx_str : str
        f(x) as SymPy-parseable string.
    params_per_method : dict[method_key → params], optional
        Per-method parameter overrides. Unspecified methods use auto-defaults.
    n_iter : int
        Iteration rows per sheet (default 25).
    eq_label : str
        Human-readable equation for all footers.
    methods : list[str], optional
        Subset of method keys to include. Defaults to ALL_METHODS.

    Returns
    -------
    bytes
        Raw .xlsx content with one sheet per method.
    """
    selected = methods or ALL_METHODS
    params_per_method = params_per_method or {}

    wb = Workbook()
    wb.remove(wb.active)

    for method_key in selected:
        if method_key not in TEMPLATE_REGISTRY:
            continue
        method_params = params_per_method.get(method_key)
        _build_sheet(wb, method_key, fx_str, method_params, n_iter, eq_label or fx_str)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Convenience: write directly to file (useful for testing)
# ---------------------------------------------------------------------------

def save_single(path: str, method_key: str, fx_str: str, **kwargs) -> None:
    data = generate_single(method_key, fx_str, **kwargs)
    with open(path, "wb") as f:
        f.write(data)


def save_all(path: str, fx_str: str, **kwargs) -> None:
    data = generate_all(fx_str, **kwargs)
    with open(path, "wb") as f:
        f.write(data)
