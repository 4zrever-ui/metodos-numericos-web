"""
excel_styles.py
---------------
Centralizes ALL visual styling for the generated Excel workbook.

Provides:
- Color palettes per method (header, light1, light2, footer_bg, footer_text_label)
- Factory functions for openpyxl Font, PatternFill, Border, Alignment objects
- Helper to apply a full style package to a cell
- Row-height constants
"""

from __future__ import annotations
from dataclasses import dataclass
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, Color
)


# ---------------------------------------------------------------------------
# Color Palette Definitions (exact hex values from amburger.xlsx analysis)
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class MethodPalette:
    header: str        # Column header row background
    light1: str        # Odd data rows (k=1,3,5…)
    light2: str        # Even data rows (k=0,2,4…) — same as header
    footer_bg: str     # Merged title/footer row background
    label_color: str   # Parameter label text color


PALETTES: dict[str, MethodPalette] = {
    "biseccion": MethodPalette(
        header="D6EAF8", light1="D6EAF8", light2="EBF5FB",
        footer_bg="1A5276", label_color="2E86C1"
    ),
    "regula_falsi": MethodPalette(
        header="D5F5E3", light1="D5F5E3", light2="EAFAF1",
        footer_bg="147F57", label_color="147F57"
    ),
    "punto_fijo": MethodPalette(
        header="E8DAEF", light1="E8DAEF", light2="F4EBFD",
        footer_bg="6C3483", label_color="6C3483"
    ),
    "aitken": MethodPalette(
        header="D1F2EB", light1="D1F2EB", light2="E8F8F5",
        footer_bg="147F57", label_color="147F57"
    ),
    "steffensen": MethodPalette(
        header="D1F2EB", light1="D1F2EB", light2="D3EAFD",
        footer_bg="147F57", label_color="147F57"
    ),
    "newton_raphson": MethodPalette(
        header="D6DBFF", light1="D6DBFF", light2="EEF0FF",
        footer_bg="1F3A93", label_color="3455DB"
    ),
    "newton_modificado": MethodPalette(
        header="E74C3C", light1="FDEDEC", light2="FADBD8",
        footer_bg="922B21", label_color="922B21"
    ),
    "newton_2do_orden": MethodPalette(
        header="D5F5E3", light1="D5F5E3", light2="EAFAF1",
        footer_bg="0D553B", label_color="0D553B"
    ),
    "chebyshev": MethodPalette(
        header="D2E4F8", light1="D2E4F8", light2="D2E4F8",
        footer_bg="174969", label_color="174969"
    ),
    "halley": MethodPalette(
        header="E8F5D6", light1="E8F5D6", light2="F0FAE8",
        footer_bg="4D6A1F", label_color="4D6A1F"
    ),
    "super_halley": MethodPalette(
        header="FADBD8", light1="FADBD8", light2="FEF5F5",
        footer_bg="7B241C", label_color="7B241C"
    ),
    "ostrowsky": MethodPalette(
        header="E8DAEF", light1="E8DAEF", light2="F5EEF8",
        footer_bg="4A235A", label_color="4A235A"
    ),
    "secante": MethodPalette(
        header="D1F2EB", light1="D1F2EB", light2="E8F8F5",
        footer_bg="0D553B", label_color="0D553B"
    ),
    "von_mises": MethodPalette(
        header="D6EAF8", light1="D6EAF8", light2="EBF5FB",
        footer_bg="1A5276", label_color="2E86C1"
    ),
}

# ---------------------------------------------------------------------------
# Text colors (universal)
# ---------------------------------------------------------------------------
DATA_TEXT_COLOR = "1A1A2E"       # All data cells
FOOTER_TEXT_COLOR = "FFFFFF"     # White text on dark footer background
HEADER_TEXT_COLOR = "1A1A2E"     # Column header labels

# ---------------------------------------------------------------------------
# Row heights (points)
# ---------------------------------------------------------------------------
ROW_HEIGHT_DEFAULT = 15.0
ROW_HEIGHT_HEADER = 17.25
ROW_HEIGHT_FOOTER = 17.25

# ---------------------------------------------------------------------------
# Border helpers
# ---------------------------------------------------------------------------
_THIN = Side(style="thin", color="000000")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
NO_BORDER = Border()


def thin_border() -> Border:
    return THIN_BORDER


# ---------------------------------------------------------------------------
# Style factory functions
# ---------------------------------------------------------------------------

def header_font() -> Font:
    return Font(name="Calibri", size=11, color=HEADER_TEXT_COLOR, bold=False)


def data_font() -> Font:
    return Font(name="Calibri", size=11, color=DATA_TEXT_COLOR, bold=False)


def footer_font() -> Font:
    return Font(name="Calibri", size=11, color=FOOTER_TEXT_COLOR, bold=True)


def label_font(color: str) -> Font:
    return Font(name="Calibri", size=11, color=color, bold=False)


def solid_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", start_color=hex_color, end_color=hex_color)


def default_alignment() -> Alignment:
    return Alignment(vertical="center")


def center_alignment() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# High-level apply helpers
# ---------------------------------------------------------------------------

def apply_header_style(cell, palette: MethodPalette) -> None:
    """Style for column-header row (row 1)."""
    cell.font = header_font()
    cell.fill = solid_fill(palette.header)
    cell.border = thin_border()
    cell.alignment = center_alignment()


def apply_data_style(cell, row_index: int, palette: MethodPalette) -> None:
    """
    Style for data rows. row_index is the 0-based iteration index (k value).
    Even k → light2 (lighter), Odd k → light1 (same as header).
    """
    bg = palette.light2 if row_index % 2 == 0 else palette.light1
    cell.font = data_font()
    cell.fill = solid_fill(bg)
    cell.border = thin_border()
    cell.alignment = default_alignment()


def apply_footer_style(cell, palette: MethodPalette) -> None:
    """Style for merged title/footer row."""
    cell.font = footer_font()
    cell.fill = solid_fill(palette.footer_bg)
    cell.border = thin_border()
    cell.alignment = center_alignment()


def apply_label_style(cell, palette: MethodPalette) -> None:
    """Style for parameter label cells (Ecuación:, Intervalo:, etc.)."""
    cell.font = label_font(palette.label_color)
    cell.border = thin_border()
    cell.alignment = default_alignment()


def apply_param_value_style(cell, palette: MethodPalette) -> None:
    """Style for parameter value cells (the value next to a label)."""
    cell.font = data_font()
    cell.border = thin_border()
    cell.alignment = default_alignment()


def style_full_row(ws, excel_row: int, num_cols: int,
                   k_index: int, palette: MethodPalette) -> None:
    """Apply data style to an entire iteration row."""
    for col in range(1, num_cols + 1):
        apply_data_style(ws.cell(excel_row, col), k_index, palette)


def style_header_row(ws, excel_row: int, num_cols: int,
                     palette: MethodPalette) -> None:
    """Apply header style to an entire header row."""
    for col in range(1, num_cols + 1):
        apply_header_style(ws.cell(excel_row, col), palette)
