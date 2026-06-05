"""
excel_templates.py
------------------
One template-builder class per numerical method.

Each class receives:
  - ws          : openpyxl Worksheet (already named, added to wb externally)
  - fx_str      : f(x) as sympy-parseable string, e.g. "x**2 - 2"
  - fpx_str     : f'(x) as sympy-parseable string (None for non-derivative methods)
  - fppx_str    : f''(x) as sympy-parseable string (None when not needed)
  - params      : dict with method-specific initial values:
                    a0, b0       → interval methods
                    x0           → point methods
                    x0, x1       → secant
                    g_str        → punto fijo / aitken / steffensen
  - n_iter      : number of iteration rows to generate (default 25)
  - eq_label    : human-readable equation string for footer, e.g. "f(x) = x² - 2"

Each class writes:
  1. Column header row (row 1)
  2. Iteration rows with REAL Excel formulas (rows 2 … n_iter+1)
  3. Merged footer title row
  4. Parameter info rows

All column widths are set at the end.
Styles are applied via excel_styles module.
Formulas are generated via sympy_to_excel module.
"""

from __future__ import annotations
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

from excel_styles import (
    PALETTES, MethodPalette,
    apply_header_style, apply_data_style, apply_footer_style,
    apply_label_style, apply_param_value_style,
    style_full_row, style_header_row,
    ROW_HEIGHT_DEFAULT, ROW_HEIGHT_HEADER, ROW_HEIGHT_FOOTER,
    data_font, solid_fill, thin_border, default_alignment,
)
from sympy_to_excel import (
    to_excel_formula, derivative_to_excel,
    evaluate_derivative_at, expr_to_sympy,
)

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _set_col_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header(ws, headers: list[str], palette: MethodPalette) -> None:
    ws.row_dimensions[1].height = ROW_HEIGHT_HEADER
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        apply_header_style(c, palette)


def _write_footer(ws, footer_row: int, merge_range: str,
                  title_text: str, palette: MethodPalette,
                  param_rows: list[list[tuple[int, str, bool]]],
                  ) -> None:
    """
    Write merged title row and parameter rows below it.

    param_rows: list of rows, each row is a list of (col, value, is_label) tuples.
    """
    ws.row_dimensions[footer_row].height = ROW_HEIGHT_FOOTER
    ws.merge_cells(merge_range)
    c = ws.cell(footer_row, 1, title_text)
    apply_footer_style(c, palette)

    for offset, row_def in enumerate(param_rows, 1):
        r = footer_row + offset
        ws.row_dimensions[r].height = ROW_HEIGHT_DEFAULT
        for (col, val, is_label) in row_def:
            cell = ws.cell(r, col, val)
            if is_label:
                apply_label_style(cell, palette)
            else:
                apply_param_value_style(cell, palette)


def _data_rows(ws, first_row: int, last_row: int,
               palette: MethodPalette, num_cols: int) -> None:
    """Pre-apply data background to every iteration row."""
    for r in range(first_row, last_row + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT_DEFAULT
        k = r - first_row   # k-index for alternating colors
        for col in range(1, num_cols + 1):
            apply_data_style(ws.cell(r, col), k, palette)


# ---------------------------------------------------------------------------
# 1. BISECCIÓN
# ---------------------------------------------------------------------------

class BiseccionTemplate:
    METHOD_KEY = "biseccion"
    HEADERS = ["k", "a", "c=(a+b)/2", "b", "f(a)", "f(c)", "f(b)", "f(a)·f(c)", "Error %", "Convergencia"]
    N_COLS = 10

    def build(self, ws, fx_str: str, params: dict, n_iter: int = 25, eq_label: str = "") -> None:
        p = PALETTES[self.METHOD_KEY]
        a0 = params.get("a0", 1)
        b0 = params.get("b0", 2)

        _write_header(ws, self.HEADERS, p)
        _data_rows(ws, 2, n_iter + 1, p, self.N_COLS)

        fx = lambda ref: to_excel_formula(fx_str, ref)

        for i in range(n_iter):
            r = i + 2  # Excel row
            k = i
            ws.cell(r, 1, k if k == 0 else f"={r-2}")
            if k == 0:
                ws.cell(r, 1, 0)
                ws.cell(r, 2, a0)
                ws.cell(r, 4, b0)
            else:
                prev_E = f"E{r-1}"
                prev_F = f"F{r-1}"
                prev_B = f"B{r-1}"
                prev_C = f"C{r-1}"
                prev_D = f"D{r-1}"
                ws.cell(r, 1, k)
                ws.cell(r, 2, f"=IF({prev_E}*{prev_F}<0,{prev_B},{prev_C})")
                ws.cell(r, 4, f"=IF({prev_E}*{prev_F}<0,{prev_C},{prev_D})")

            bref, dref = f"B{r}", f"D{r}"
            ws.cell(r, 3, f"=({bref}+{dref})/2")
            ws.cell(r, 5, f"={fx(bref)}")
            ws.cell(r, 6, f"={fx(f'C{r}')}")
            ws.cell(r, 7, f"={fx(dref)}")
            ws.cell(r, 8, f"=E{r}*F{r}")

            if k == 0:
                ws.cell(r, 9, "")
                ws.cell(r, 10, "")
            else:
                ws.cell(r, 9, f"=ABS((C{r}-C{r-1})/C{r})*100")
                ws.cell(r, 10, f'=IF(I{r}<0.00001,"SI","NO")')

        footer_row = n_iter + 2
        _write_footer(
            ws, footer_row,
            f"A{footer_row}:J{footer_row}",
            f"MÉTODO DE BISECCIÓN  —  {eq_label}", p,
            [[
                (1, "Ecuación:", True), (2, eq_label, False),
                (4, "Intervalo inicial:", True), (5, f"[a₀, b₀] = [{a0}, {b0}]", False),
                (7, "Tolerancia:", True), (8, "ε < 1×10⁻⁵ %", False),
            ]]
        )
        _set_col_widths(ws, [5, 14, 14, 14, 14, 14, 14, 14, 12, 14])


# ---------------------------------------------------------------------------
# 2. REGULA FALSI
# ---------------------------------------------------------------------------

class RegulaFalsiTemplate:
    METHOD_KEY = "regula_falsi"
    HEADERS = ["k", "a", "c (R.F.)", "b", "f(a)", "f(c)", "f(b)", "f(a)·f(c)", "Error %", "Convergencia"]
    N_COLS = 10

    def build(self, ws, fx_str: str, params: dict, n_iter: int = 25, eq_label: str = "") -> None:
        p = PALETTES[self.METHOD_KEY]
        a0 = params.get("a0", 1)
        b0 = params.get("b0", 2)

        _write_header(ws, self.HEADERS, p)
        _data_rows(ws, 2, n_iter + 1, p, self.N_COLS)

        fx = lambda ref: to_excel_formula(fx_str, ref)

        for i in range(n_iter):
            r = i + 2
            k = i
            if k == 0:
                ws.cell(r, 1, 0)
                ws.cell(r, 2, a0)
                ws.cell(r, 4, b0)
            else:
                prev_E, prev_F = f"E{r-1}", f"F{r-1}"
                prev_B, prev_C, prev_D = f"B{r-1}", f"C{r-1}", f"D{r-1}"
                ws.cell(r, 1, k)
                ws.cell(r, 2, f"=IF({prev_E}*{prev_F}<0,{prev_B},{prev_C})")
                ws.cell(r, 4, f"=IF({prev_E}*{prev_F}<0,{prev_C},{prev_D})")

            bref, dref = f"B{r}", f"D{r}"
            ws.cell(r, 5, f"={fx(bref)}")
            ws.cell(r, 7, f"={fx(dref)}")
            # Regula Falsi formula: c = a - f(a)*(b-a)/(f(b)-f(a))
            ws.cell(r, 3, f"={bref}-(E{r}*({dref}-{bref}))/(G{r}-E{r})")
            ws.cell(r, 6, f"={fx(f'C{r}')}")
            ws.cell(r, 8, f"=E{r}*F{r}")

            if k == 0:
                ws.cell(r, 9, "")
                ws.cell(r, 10, "")
            else:
                ws.cell(r, 9, f"=ABS((C{r}-C{r-1})/C{r})*100")
                ws.cell(r, 10, f'=IF(I{r}<0.00001,"SI","NO")')

        footer_row = n_iter + 2
        _write_footer(
            ws, footer_row,
            f"A{footer_row}:J{footer_row}",
            f"MÉTODO DE REGULA FALSI  —  {eq_label}", p,
            [[
                (1, "Ecuación:", True), (2, eq_label, False),
                (4, "Intervalo inicial:", True), (5, f"[a₀, b₀] = [{a0}, {b0}]", False),
                (7, "Tolerancia:", True), (8, "ε < 1×10⁻⁵ %", False),
            ]]
        )
        _set_col_widths(ws, [5, 14, 14, 14, 14, 14, 14, 14, 12, 14])


# ---------------------------------------------------------------------------
# 3. PUNTO FIJO
# ---------------------------------------------------------------------------

class PuntoFijoTemplate:
    METHOD_KEY = "punto_fijo"
    HEADERS = ["k", "xₖ", "g(xₖ)", "Error %", "Convergencia"]
    N_COLS = 5

    def build(self, ws, fx_str: str, params: dict, n_iter: int = 25, eq_label: str = "") -> None:
        p = PALETTES[self.METHOD_KEY]
        x0 = params.get("x0", 1)
        g_str = params.get("g_str", "")

        _write_header(ws, self.HEADERS, p)
        _data_rows(ws, 2, n_iter + 1, p, self.N_COLS)

        gx = lambda ref: to_excel_formula(g_str, ref)

        for i in range(n_iter):
            r = i + 2
            k = i
            ws.cell(r, 1, k)
            if k == 0:
                ws.cell(r, 2, x0)
                ws.cell(r, 4, "")
                ws.cell(r, 5, "")
            else:
                ws.cell(r, 2, f"=C{r-1}")
                ws.cell(r, 4, f"=ABS((C{r}-B{r})/C{r})*100")
                ws.cell(r, 5, f'=IF(D{r}<0.00001,"SI","NO")')

            ws.cell(r, 3, f"={gx(f'B{r}')}")

        footer_row = n_iter + 2
        g_display = params.get("g_display", g_str)
        _write_footer(
            ws, footer_row,
            f"A{footer_row}:E{footer_row}",
            f"MÉTODO DE PUNTO FIJO  —  {eq_label}", p,
            [[
                (1, "Ecuación:", True), (2, eq_label, False),
                (4, "g(x):", True), (5, g_display, False),
            ]]
        )
        _set_col_widths(ws, [5, 16, 20, 12, 14])


# ---------------------------------------------------------------------------
# 4. AITKEN (Δ²)
# ---------------------------------------------------------------------------

class AitkenTemplate:
    METHOD_KEY = "aitken"
    HEADERS_MAIN = ["k", "p̂ₖ (Aitken)", "Error %", "Convergencia"]
    HEADERS_AUX = ["k", "xₖ", "g(xₖ)", "Error %", "Convergencia"]
    N_COLS_MAIN = 4

    def build(self, ws, fx_str: str, params: dict, n_iter: int = 6, eq_label: str = "") -> None:
        p = PALETTES[self.METHOD_KEY]
        x0 = params.get("x0", 1)
        g_str = params.get("g_str", "")
        # Auxiliary table starts at column I (9), row 1
        AUX_COL = 9   # I

        # -- Auxiliary Punto Fijo table (cols I–M) --
        p_pf = PALETTES["punto_fijo"]

        # Row 1: title banner for PF aux table
        ws.merge_cells("I1:M1")
        pf_title_cell = ws.cell(1, AUX_COL, "PUNTO FIJO  —  base de Aitken")
        apply_footer_style(pf_title_cell, p_pf)
        ws.row_dimensions[1].height = ROW_HEIGHT_FOOTER

        # Row 2: column headers
        aux_headers = ["k", "xₖ", "g(xₖ)", "Error %", "Conv"]
        for ci, h in enumerate(aux_headers, AUX_COL):
            c = ws.cell(2, ci, h)
            apply_header_style(c, p_pf)

        gx = lambda ref: to_excel_formula(g_str, ref)
        aux_rows = n_iter + 4   # need extra rows for Δ² lookback

        # Variante A sliding-window bands of 3 rows
        GROUP_COLORS = [p_pf.light1, p_pf.light2]

        for i in range(aux_rows):
            r = i + 3  # data starts at row 3 (title=1, header=2)
            k = i
            ws.cell(r, AUX_COL, k)
            if k == 0:
                ws.cell(r, AUX_COL + 1, x0)
                ws.cell(r, AUX_COL + 3, "")
                ws.cell(r, AUX_COL + 4, "")
            else:
                ws.cell(r, AUX_COL + 1, f"={get_column_letter(AUX_COL+2)}{r-1}")
                ws.cell(r, AUX_COL + 3, f"=ABS(({get_column_letter(AUX_COL+2)}{r}-{get_column_letter(AUX_COL+1)}{r})/{get_column_letter(AUX_COL+2)}{r})*100")
                ws.cell(r, AUX_COL + 4, f'=IF({get_column_letter(AUX_COL+3)}{r}<0.00001,"SI","NO")')
            aux_x = f"{get_column_letter(AUX_COL+1)}{r}"
            ws.cell(r, AUX_COL + 2, f"={gx(aux_x)}")
            # Variante A sliding-window color:
            # x̂ₖ uses PF rows k, k+1, k+2. Color in bands of 3 rows so
            # x̂₀'s source rows (0,1,2) share one color, x̂₃'s (3,4,5) share another, etc.
            group_idx = (i // 3) % 2
            row_color = GROUP_COLORS[group_idx % 2]
            for ci in range(AUX_COL, AUX_COL + 5):
                cell = ws.cell(r, ci)
                cell.font = data_font()
                cell.fill = solid_fill(row_color)
                cell.border = thin_border()
                cell.alignment = default_alignment()

        # -- Aitken main table (cols A–D) --
        _write_header(ws, self.HEADERS_MAIN, p)
        # Aitken Δ² formula uses K column (AUX_COL+2 = K).
        # PF data now starts at row 3 (row 1=title, row 2=header).
        K = get_column_letter(AUX_COL + 2)

        for i in range(n_iter):
            r = i + 2   # Aitken result row (cols A–D)
            k = i
            pf_r = i + 3  # corresponding PF data row in col K
            ws.cell(r, 1, k)
            # p̂ₖ = K[pf_r] - (K[pf_r+1] - K[pf_r])² / (K[pf_r+2] - 2·K[pf_r+1] + K[pf_r])
            kr, kr1, kr2 = f"{K}{pf_r}", f"{K}{pf_r+1}", f"{K}{pf_r+2}"
            ws.cell(r, 2, f"={kr}-(({kr1}-{kr})^2)/({kr2}-2*{kr1}+{kr})")

            if k == 0:
                ws.cell(r, 3, "")
                ws.cell(r, 4, "")
            else:
                ws.cell(r, 3, f"=ABS((B{r}-B{r-1})/B{r})*100")
                ws.cell(r, 4, f'=IF(C{r}<0.00001,"SI","NO")')
            for ci in range(1, 5):
                apply_data_style(ws.cell(r, ci), k, p)

        footer_row = n_iter + 2
        _write_footer(
            ws, footer_row,
            f"A{footer_row}:D{footer_row}",
            f"MÉTODO DE AITKEN (Δ²)  —  {eq_label}", p,
            [[(1, "Ecuación:", True), (2, eq_label, False), (4, "g(x):", True)]]
        )
        _set_col_widths(ws, [5, 20, 12, 14, 5, 5, 5, 5, 5, 16, 20, 12, 14])


# ---------------------------------------------------------------------------
# 5. STEFFENSEN
# ---------------------------------------------------------------------------

class SteffensenTemplate:
    METHOD_KEY = "steffensen"

    def build(self, ws, fx_str: str, params: dict, n_iter: int = 4, eq_label: str = "") -> None:
        """
        Three visible layers, faithful to amburger.xlsx structure:
          Layer 1 — Steffensen  (cols A–D, rows 1–5)
          Layer 2 — Aitken aux  (cols I–L, rows 6–13)
          Layer 3 — Punto Fijo  (cols I–M, rows 17–29)
        All columns visible. Formulas reference exactly as in original.
        """
        p = PALETTES[self.METHOD_KEY]
        p_ait = PALETTES["aitken"]
        p_pf  = PALETTES["punto_fijo"]

        x0    = params.get("x0", 1)
        g_str = params.get("g_str", "")
        gx    = lambda ref: to_excel_formula(g_str, ref)

        # Column indices (1-based)
        COL_I = 9   # I  — Aitken/PF table start
        COL_J = 10  # J  — Aitken values / PF xₖ
        COL_K = 11  # K  — PF g(xₖ)
        COL_L = 12
        COL_M = 13

        J = get_column_letter(COL_J)   # "J"
        K = get_column_letter(COL_K)   # "K"
        L = get_column_letter(COL_L)   # "L"

        # ── ROW ANCHORS (match amburger.xlsx) ────────────────────────
        STF_HDR  = 1    # Steffensen header
        STF_DATA = 2    # Steffensen data rows (2–3)
        STF_FTR  = 4    # Steffensen footer / title
        STF_LBL  = 5    # Ecuación / params row

        AIT_HDR  = 6    # Aitken header
        AIT_DATA = 7    # Aitken data rows (7–11)
        AIT_FTR  = 12   # Aitken footer / title
        AIT_LBL  = 13   # Ecuación row

        PF_LBL   = 16   # "MÉTODO DE PUNTO FIJO" title row (blank separator)
        PF_LBL2  = 17   # g(x) label row
        PF_HDR   = 18   # PF header
        PF_DATA  = 19   # PF data rows (19–29)

        # ════════════════════════════════════════════════════════════
        # LAYER 1 — STEFFENSEN  (cols A–D)
        # ════════════════════════════════════════════════════════════

        # Header row
        for ci, h in enumerate(["k", "p̂ₖ  (Steffensen)", "Error %", "Convergencia"], 1):
            apply_header_style(ws.cell(STF_HDR, ci, h), p)

        # Data rows: 2 iterations (k=0,1). References J col of Aitken table.
        for i in range(2):
            r  = STF_DATA + i
            jr  = f"{J}{AIT_DATA + i}"
            jr1 = f"{J}{AIT_DATA + i + 1}"
            jr2 = f"{J}{AIT_DATA + i + 2}"
            ws.cell(r, 1, i)
            ws.cell(r, 2, f"={jr}-(({jr1}-{jr})^2)/({jr2}-2*{jr1}+{jr})")
            if i == 0:
                ws.cell(r, 3, "")
                ws.cell(r, 4, "")
            else:
                ws.cell(r, 3, f"=ABS((B{r}-B{r-1})/B{r})*100")
                ws.cell(r, 4, f'=IF(C{r}<0.00001,"SI","NO")')
            for ci in range(1, 5):
                apply_data_style(ws.cell(r, ci), i, p)

        # Footer / title row (merged A–D)
        ws.merge_cells(f"A{STF_FTR}:D{STF_FTR}")
        title_cell = ws.cell(STF_FTR, 1, f"MÉTODO DE STEFFENSEN  —  {eq_label}")
        apply_footer_style(title_cell, p)
        ws.row_dimensions[STF_FTR].height = ROW_HEIGHT_FOOTER

        # Param label row
        for ci, val, is_label in [
            (1, "Ecuación:", True), (2, eq_label, False),
            (4, "Base: Aitken (Δ² sobre Punto Fijo)", True),
        ]:
            c = ws.cell(STF_LBL, ci, val)
            if is_label:
                apply_label_style(c, p)
            else:
                apply_param_value_style(c, p)

        # ════════════════════════════════════════════════════════════
        # LAYER 2 — AITKEN AUXILIAR  (cols I–L, rows 6–13)
        # ════════════════════════════════════════════════════════════

        # Header
        for ci, h in enumerate(["k", "x̂ₖ  (Aitken Δ²)", "Error %", "Conv"], COL_I):
            apply_header_style(ws.cell(AIT_HDR, ci, h), p_ait)

        # Data: 4 Aitken rows (k=0..3), references K col of PF table
        for i in range(4):
            r    = AIT_DATA + i
            pf_r = PF_DATA + i          # K col row in PF table
            kr   = f"{K}{pf_r}"
            kr1  = f"{K}{pf_r + 1}"
            kr2  = f"{K}{pf_r + 2}"
            ws.cell(r, COL_I, i)
            ws.cell(r, COL_J, f"={kr}-(({kr1}-{kr})^2)/({kr2}-2*{kr1}+{kr})")
            if i == 0:
                ws.cell(r, COL_K, "")
                ws.cell(r, COL_L, "")
            else:
                ws.cell(r, COL_K, f"=ABS(({J}{r}-{J}{r-1})/{J}{r})*100")
                ws.cell(r, COL_L, f'=IF({K}{r}<0.00001,"SI","NO")')
            for ci in range(COL_I, COL_L + 1):
                apply_data_style(ws.cell(r, ci), i, p_ait)

        # Aitken footer / title
        ws.merge_cells(f"I{AIT_FTR}:L{AIT_FTR}")
        ait_title = ws.cell(AIT_FTR, COL_I, f"MÉTODO DE AITKEN (Δ²)  —  {eq_label}")
        apply_footer_style(ait_title, p_ait)
        ws.row_dimensions[AIT_FTR].height = ROW_HEIGHT_FOOTER

        # Aitken param labels
        for ci, val, is_label in [
            (COL_I, "Ecuación:", True), (COL_J, eq_label, False),
            (COL_L, "Base: Punto Fijo", True),
        ]:
            c = ws.cell(AIT_LBL, ci, val)
            if is_label:
                apply_label_style(c, p_ait)
            else:
                apply_param_value_style(c, p_ait)

        # ════════════════════════════════════════════════════════════
        # LAYER 3 — PUNTO FIJO BASE  (cols I–M, rows 16–29)
        # ════════════════════════════════════════════════════════════

        # Title row (separator)
        ws.merge_cells(f"I{PF_LBL}:M{PF_LBL}")
        pf_title = ws.cell(PF_LBL, COL_I, f"MÉTODO DE PUNTO FIJO  —  {eq_label}")
        apply_footer_style(pf_title, p_pf)
        ws.row_dimensions[PF_LBL].height = ROW_HEIGHT_FOOTER

        # g(x) label row
        for ci, val, is_label in [
            (COL_I, "Ecuación:", True), (COL_J, eq_label, False),
            (COL_L, "g(x):", True), (COL_M, f"g(x) = {g_str}", False),
        ]:
            c = ws.cell(PF_LBL2, ci, val)
            if is_label:
                apply_label_style(c, p_pf)
            else:
                apply_param_value_style(c, p_pf)

        # PF header
        for ci, h in enumerate(["k", "xₖ", "g(xₖ)", "Error %", "Convergencia"], COL_I):
            apply_header_style(ws.cell(PF_HDR, ci, h), p_pf)

        # PF data: 10 rows (k=0..9) — enough to supply Aitken's lookback
        pf_rows = 10
        X = get_column_letter(COL_J)  # "J"
        G = get_column_letter(COL_K)  # "K" = g(xₖ)
        Er = get_column_letter(COL_L)
        Cv = get_column_letter(COL_M)

        for i in range(pf_rows):
            r = PF_DATA + i
            ws.cell(r, COL_I, i)
            if i == 0:
                ws.cell(r, COL_J, x0)
                ws.cell(r, COL_L, "")
                ws.cell(r, COL_M, "")
            else:
                ws.cell(r, COL_J, f"={G}{r - 1}")
                ws.cell(r, COL_L, f"=ABS(({G}{r}-{X}{r})/{G}{r})*100")
                ws.cell(r, COL_M, f'=IF({Er}{r}<0.00001,"SI","NO")')
            xk_ref = f"{X}{r}"
            ws.cell(r, COL_K, f"={gx(xk_ref)}")
            for ci in range(COL_I, COL_M + 1):
                apply_data_style(ws.cell(r, ci), i, p_pf)

        # ── Column widths (all visible) ───────────────────────────
        col_widths = {
            "A": 5, "B": 22, "C": 12, "D": 14,
            "I": 5, "J": 22, "K": 22, "L": 12, "M": 14,
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
            ws.column_dimensions[col_letter].hidden = False  # ensure visible


# ---------------------------------------------------------------------------
# 6. NEWTON-RAPHSON
# ---------------------------------------------------------------------------

class NewtonRaphsonTemplate:
    METHOD_KEY = "newton_raphson"
    HEADERS = ["k", "xₖ", "f(xₖ)", "f'(xₖ)", "xₖ₊₁", "Error %", "Convergencia"]
    N_COLS = 7

    def build(self, ws, fx_str: str, params: dict, n_iter: int = 25, eq_label: str = "",
              fpx_str: str = "") -> None:
        p = PALETTES[self.METHOD_KEY]
        x0 = params.get("x0", 1)

        _write_header(ws, self.HEADERS, p)
        _data_rows(ws, 2, n_iter + 1, p, self.N_COLS)

        fx  = lambda ref: to_excel_formula(fx_str,  ref)
        fpx = lambda ref: to_excel_formula(fpx_str, ref) if fpx_str else derivative_to_excel(fx_str, ref, 1)

        for i in range(n_iter):
            r = i + 2
            k = i
            ws.cell(r, 1, k)
            if k == 0:
                ws.cell(r, 2, x0)
            else:
                ws.cell(r, 2, f"=E{r-1}")

            bref = f"B{r}"
            ws.cell(r, 3, f"={fx(bref)}")
            ws.cell(r, 4, f"={fpx(bref)}")
            ws.cell(r, 5, f"={bref}-(C{r}/D{r})")
            ws.cell(r, 6, f"=ABS((E{r}-{bref})/E{r})*100")
            ws.cell(r, 7, f'=IF(F{r}<0.00001,"SI","NO")')

        footer_row = n_iter + 2
        _write_footer(
            ws, footer_row, f"A{footer_row}:G{footer_row}",
            f"MÉTODO DE NEWTON-RAPHSON  —  {eq_label}", p,
            [[
                (1, "Ecuación:", True), (2, eq_label, False),
                (4, "Fórmula:", True), (5, "xₙ₊₁ = xₙ − f(xₙ)/f'(xₙ)", False),
            ]]
        )
        _set_col_widths(ws, [5, 16, 16, 16, 16, 12, 14])


# ---------------------------------------------------------------------------
# 7. NEWTON MODIFICADO
# ---------------------------------------------------------------------------

class NewtonModificadoTemplate:
    METHOD_KEY = "newton_modificado"
    HEADERS = ["k", "xₖ", "f(xₖ)", "f'(xₖ)", "f''(xₖ)", "xₖ₊₁", "Error %", "Convergencia"]
    N_COLS = 8

    def build(self, ws, fx_str: str, params: dict, n_iter: int = 25, eq_label: str = "",
              fpx_str: str = "", fppx_str: str = "") -> None:
        p = PALETTES[self.METHOD_KEY]
        x0 = params.get("x0", 1)

        _write_header(ws, self.HEADERS, p)
        _data_rows(ws, 2, n_iter + 1, p, self.N_COLS)

        fx   = lambda ref: to_excel_formula(fx_str,   ref)
        fpx  = lambda ref: to_excel_formula(fpx_str,  ref) if fpx_str  else derivative_to_excel(fx_str, ref, 1)

        # f''(x) stored as constant in $B$9 (row 9 aux cell)
        fpp_val = evaluate_derivative_at(fx_str, x0, order=2)
        if fpp_val == int(fpp_val):
            fpp_val = int(fpp_val)
        aux_row = n_iter + 10
        ws.cell(aux_row, 2, fpp_val)  # $B$aux_row
        fpp_ref = f"$B${aux_row}"

        for i in range(n_iter):
            r = i + 2
            k = i
            ws.cell(r, 1, k)
            if k == 0:
                ws.cell(r, 2, x0)
            else:
                ws.cell(r, 2, f"=F{r-1}")

            bref = f"B{r}"
            ws.cell(r, 3, f"={fx(bref)}")
            ws.cell(r, 4, f"={fpx(bref)}")
            ws.cell(r, 5, f"={fpp_ref}")
            # xₙ₊₁ = xₙ - f·f' / (f'² - f·f'')
            ws.cell(r, 6, f"={bref}-((C{r}*D{r})/(D{r}^2-C{r}*E{r}))")
            ws.cell(r, 7, f"=ABS((F{r}-{bref})/F{r})*100")
            ws.cell(r, 8, f'=IF(G{r}<0.00001,"SI","NO")')

        footer_row = n_iter + 2
        _write_footer(
            ws, footer_row, f"A{footer_row}:H{footer_row}",
            f"MÉTODO DE NEWTON MODIFICADO  —  {eq_label}", p,
            [[
                (1, "Ecuación:", True), (2, eq_label, False),
                (5, "f''(x₀):", True), (6, str(fpp_val), False),
            ]]
        )
        _set_col_widths(ws, [5, 16, 16, 16, 14, 16, 12, 14])


# ---------------------------------------------------------------------------
# 8. NEWTON 2DO ORDEN
# ---------------------------------------------------------------------------

class Newton2doOrdenTemplate:
    METHOD_KEY = "newton_2do_orden"
    HEADERS = ["k", "xₖ", "f(xₖ)", "f'(xₖ)", "f''(xₖ)", "xₖ₊₁", "Error %", "Convergencia"]
    N_COLS = 8

    def build(self, ws, fx_str: str, params: dict, n_iter: int = 25, eq_label: str = "",
              fpx_str: str = "", fppx_str: str = "") -> None:
        p = PALETTES[self.METHOD_KEY]
        x0 = params.get("x0", 1)

        _write_header(ws, self.HEADERS, p)
        _data_rows(ws, 2, n_iter + 1, p, self.N_COLS)

        fx  = lambda ref: to_excel_formula(fx_str,  ref)
        fpx = lambda ref: to_excel_formula(fpx_str, ref) if fpx_str else derivative_to_excel(fx_str, ref, 1)
        fpp_val = evaluate_derivative_at(fx_str, x0, order=2)
        if fpp_val == int(fpp_val):
            fpp_val = int(fpp_val)

        for i in range(n_iter):
            r = i + 2
            k = i
            ws.cell(r, 1, k)
            if k == 0:
                ws.cell(r, 2, x0)
            else:
                ws.cell(r, 2, f"=F{r-1}")

            bref = f"B{r}"
            ws.cell(r, 3, f"={fx(bref)}")
            ws.cell(r, 4, f"={fpx(bref)}")
            ws.cell(r, 5, fpp_val)
            # xₙ₊₁ = xₙ + (-f' + sqrt(f'² - 2f''·f)) / f''
            ws.cell(r, 6, f"={bref}+((-D{r}+((D{r})^2-2*E{r}*C{r})^(1/2))/E{r})")
            ws.cell(r, 7, f"=ABS((F{r}-{bref})/F{r})*100")
            ws.cell(r, 8, f'=IF(G{r}<0.00001,"SI","NO")')

        footer_row = n_iter + 2
        _write_footer(
            ws, footer_row, f"A{footer_row}:H{footer_row}",
            f"MÉTODO DE NEWTON 2DO ORDEN  —  {eq_label}", p,
            [[(1, "Ecuación:", True), (2, eq_label, False)]]
        )
        _set_col_widths(ws, [5, 16, 16, 16, 14, 16, 12, 14])


# ---------------------------------------------------------------------------
# 9. CHEBYSHEV
# ---------------------------------------------------------------------------

class ChebyshevTemplate:
    METHOD_KEY = "chebyshev"
    HEADERS = ["k", "xₖ", "f(xₖ)", "f'(xₖ)", "f''(xₖ)", "xₖ₊₁", "Error %", "Convergencia"]
    N_COLS = 8

    def build(self, ws, fx_str: str, params: dict, n_iter: int = 25, eq_label: str = "",
              fpx_str: str = "", fppx_str: str = "") -> None:
        p = PALETTES[self.METHOD_KEY]
        x0 = params.get("x0", 1)

        _write_header(ws, self.HEADERS, p)
        _data_rows(ws, 2, n_iter + 1, p, self.N_COLS)

        fx  = lambda ref: to_excel_formula(fx_str,  ref)
        fpx = lambda ref: to_excel_formula(fpx_str, ref) if fpx_str else derivative_to_excel(fx_str, ref, 1)
        fpp_val = evaluate_derivative_at(fx_str, x0, order=2)
        if fpp_val == int(fpp_val):
            fpp_val = int(fpp_val)

        # f'' stored in $B$aux
        aux_row = n_iter + 10
        ws.cell(aux_row, 2, fpp_val)
        fpp_ref = f"$B${aux_row}"

        for i in range(n_iter):
            r = i + 2
            k = i
            ws.cell(r, 1, k)
            if k == 0:
                ws.cell(r, 2, x0)
            else:
                ws.cell(r, 2, f"=F{r-1}")

            bref = f"B{r}"
            ws.cell(r, 3, f"={fx(bref)}")
            ws.cell(r, 4, f"={fpx(bref)}")
            ws.cell(r, 5, f"={fpp_ref}")
            # xₙ₊₁ = xₙ - f/f' - (f²·f'') / (2·f'³)
            ws.cell(r, 6, f"={bref}-(C{r}/D{r})-(((C{r}^2)*(E{r}))/(2*(D{r})^3))")
            ws.cell(r, 7, f"=ABS((F{r}-{bref})/F{r})*100")
            ws.cell(r, 8, f'=IF(G{r}<0.00001,"SI","NO")')

        footer_row = n_iter + 2
        _write_footer(
            ws, footer_row, f"A{footer_row}:H{footer_row}",
            f"MÉTODO DE CHEBYSHEV  —  {eq_label}", p,
            [[(1, "Ecuación:", True), (2, eq_label, False)]]
        )
        _set_col_widths(ws, [5, 16, 16, 16, 14, 16, 12, 14])


# ---------------------------------------------------------------------------
# 10. HALLEY
# ---------------------------------------------------------------------------

class HalleyTemplate:
    METHOD_KEY = "halley"
    HEADERS = ["k", "xₖ", "f(xₖ)", "f'(xₖ)", "f''(xₖ)", "xₖ₊₁", "Error %", "Convergencia"]
    N_COLS = 8

    def build(self, ws, fx_str: str, params: dict, n_iter: int = 25, eq_label: str = "",
              fpx_str: str = "", fppx_str: str = "") -> None:
        p = PALETTES[self.METHOD_KEY]
        x0 = params.get("x0", 1)

        _write_header(ws, self.HEADERS, p)
        _data_rows(ws, 2, n_iter + 1, p, self.N_COLS)

        fx  = lambda ref: to_excel_formula(fx_str,  ref)
        fpx = lambda ref: to_excel_formula(fpx_str, ref) if fpx_str else derivative_to_excel(fx_str, ref, 1)
        fpp_val = evaluate_derivative_at(fx_str, x0, order=2)
        if fpp_val == int(fpp_val):
            fpp_val = int(fpp_val)

        aux_row = n_iter + 10
        ws.cell(aux_row, 2, fpp_val)
        fpp_ref = f"$B${aux_row}"

        for i in range(n_iter):
            r = i + 2
            k = i
            ws.cell(r, 1, k)
            if k == 0:
                ws.cell(r, 2, x0)
            else:
                ws.cell(r, 2, f"=F{r-1}")

            bref = f"B{r}"
            ws.cell(r, 3, f"={fx(bref)}")
            ws.cell(r, 4, f"={fpx(bref)}")
            ws.cell(r, 5, f"={fpp_ref}")
            # xₙ₊₁ = xₙ - f / (f' - f''·f/(2·f'))
            ws.cell(r, 6, f"={bref}-(C{r}/(D{r}-(E{r}*C{r})/(2*D{r})))")
            ws.cell(r, 7, f"=ABS((F{r}-{bref})/F{r})*100")
            ws.cell(r, 8, f'=IF(G{r}<0.00001,"SI","NO")')

        footer_row = n_iter + 2
        _write_footer(
            ws, footer_row, f"A{footer_row}:H{footer_row}",
            f"MÉTODO DE HALLEY  —  {eq_label}", p,
            [[(1, "Ecuación:", True), (2, eq_label, False)]]
        )
        _set_col_widths(ws, [5, 16, 16, 16, 14, 16, 12, 14])


# ---------------------------------------------------------------------------
# 11. SUPER HALLEY
# ---------------------------------------------------------------------------

class SuperHalleyTemplate:
    METHOD_KEY = "super_halley"
    HEADERS = ["k", "xₖ", "f(xₖ)", "f'(xₖ)", "f''(xₖ)", "xₖ₊₁", "Error %", "Convergencia"]
    N_COLS = 8

    def build(self, ws, fx_str: str, params: dict, n_iter: int = 25, eq_label: str = "",
              fpx_str: str = "", fppx_str: str = "") -> None:
        p = PALETTES[self.METHOD_KEY]
        x0 = params.get("x0", 1)

        _write_header(ws, self.HEADERS, p)
        _data_rows(ws, 2, n_iter + 1, p, self.N_COLS)

        fx  = lambda ref: to_excel_formula(fx_str,  ref)
        fpx = lambda ref: to_excel_formula(fpx_str, ref) if fpx_str else derivative_to_excel(fx_str, ref, 1)
        fpp_val = evaluate_derivative_at(fx_str, x0, order=2)
        if fpp_val == int(fpp_val):
            fpp_val = int(fpp_val)

        aux_row = n_iter + 10
        ws.cell(aux_row, 2, fpp_val)
        fpp_ref = f"$B${aux_row}"

        for i in range(n_iter):
            r = i + 2
            k = i
            ws.cell(r, 1, k)
            if k == 0:
                ws.cell(r, 2, x0)
            else:
                ws.cell(r, 2, f"=F{r-1}")

            bref = f"B{r}"
            ws.cell(r, 3, f"={fx(bref)}")
            ws.cell(r, 4, f"={fpx(bref)}")
            ws.cell(r, 5, f"={fpp_ref}")
            # xₙ₊₁ = xₙ - (2f'² - f·f'') / (2(f'² - f·f'')) · (f/f')
            ws.cell(r, 6, f"={bref}-(2*(D{r}^2)-C{r}*E{r})/(2*((D{r}^2)-C{r}*E{r}))*((C{r})/(D{r}))")
            ws.cell(r, 7, f"=ABS((F{r}-{bref})/F{r})*100")
            ws.cell(r, 8, f'=IF(G{r}<0.00001,"SI","NO")')

        footer_row = n_iter + 2
        _write_footer(
            ws, footer_row, f"A{footer_row}:H{footer_row}",
            f"MÉTODO DE SUPER HALLEY  —  {eq_label}", p,
            [[(1, "Ecuación:", True), (2, eq_label, False)]]
        )
        _set_col_widths(ws, [5, 16, 16, 16, 14, 16, 12, 14])


# ---------------------------------------------------------------------------
# 12. OSTROWSKY
# ---------------------------------------------------------------------------

class OstrowskyTemplate:
    METHOD_KEY = "ostrowsky"
    HEADERS = ["k", "xₖ", "f(xₖ)", "f'(xₖ)", "f''(xₖ)", "xₖ₊₁", "Error %", "Convergencia"]
    N_COLS = 8

    def build(self, ws, fx_str: str, params: dict, n_iter: int = 25, eq_label: str = "",
              fpx_str: str = "", fppx_str: str = "") -> None:
        p = PALETTES[self.METHOD_KEY]
        x0 = params.get("x0", 1)

        # Note: Ostrowsky header is in row 2, data starts row 3
        _write_header(ws, self.HEADERS, p)
        _data_rows(ws, 3, n_iter + 2, p, self.N_COLS)  # shifted +1

        fx  = lambda ref: to_excel_formula(fx_str,  ref)
        fpx = lambda ref: to_excel_formula(fpx_str, ref) if fpx_str else derivative_to_excel(fx_str, ref, 1)
        fpp_val = evaluate_derivative_at(fx_str, x0, order=2)
        if fpp_val == int(fpp_val):
            fpp_val = int(fpp_val)

        for i in range(n_iter):
            r = i + 3  # data starts at row 3
            k = i
            ws.cell(r, 1, k)
            if k == 0:
                ws.cell(r, 2, x0)
            else:
                ws.cell(r, 2, f"=F{r-1}")

            bref = f"B{r}"
            ws.cell(r, 3, f"={fx(bref)}")
            ws.cell(r, 4, f"={fpx(bref)}")
            ws.cell(r, 5, fpp_val)  # literal value per original
            # xₙ₊₁ = xₙ - (f'/sqrt(f'²-f·f'')) · (f/f')
            ws.cell(r, 6, f"={bref}-(D{r}/((D{r}^2-C{r}*E{r})^(1/2)))*(C{r}/D{r})")
            ws.cell(r, 7, f"=ABS((F{r}-{bref})/F{r})*100")
            ws.cell(r, 8, f'=IF(G{r}<0.00001,"SI","NO")')

        footer_row = n_iter + 3
        _write_footer(
            ws, footer_row, f"A{footer_row}:H{footer_row}",
            f"MÉTODO DE OSTROWSKY  —  {eq_label}", p,
            [[(1, "Ecuación:", True), (2, eq_label, False)]]
        )
        _set_col_widths(ws, [5, 16, 16, 16, 14, 16, 12, 14])


# ---------------------------------------------------------------------------
# 13. SECANTE
# ---------------------------------------------------------------------------

class SecanteTemplate:
    METHOD_KEY = "secante"
    HEADERS = ["k", "xₖ", "f(xₖ)", "xₖ₋₁", "f(xₖ₋₁)", "xₖ₊₁", "Error %", "Convergencia"]
    N_COLS = 8

    def build(self, ws, fx_str: str, params: dict, n_iter: int = 25, eq_label: str = "") -> None:
        p = PALETTES[self.METHOD_KEY]
        x0 = params.get("x0", 1)
        x1 = params.get("x1", 2)

        _write_header(ws, self.HEADERS, p)
        _data_rows(ws, 2, n_iter + 2, p, self.N_COLS)

        fx = lambda ref: to_excel_formula(fx_str, ref)

        # k=0 row (row 2): only x0 and f(x0)
        ws.cell(2, 1, 0)
        ws.cell(2, 2, x0)
        ws.cell(2, 3, f"={fx('B2')}")
        # cols D,E,F,G,H empty for k=0
        for col in range(4, 9):
            ws.cell(2, col, "")

        # k=1 row (row 3): x1 hardcoded, x0 as D3
        ws.cell(3, 1, 1)
        ws.cell(3, 2, x1)
        ws.cell(3, 3, f"={fx('B3')}")
        ws.cell(3, 4, "=B2")
        ws.cell(3, 5, "=C2")
        ws.cell(3, 6, "=B3-(C3*(B3-D3)/(C3-E3))")
        ws.cell(3, 7, "=IFERROR(ABS((F3-B3)/F3)*100,0)")
        ws.cell(3, 8, '=IF(IFERROR(G3,0)<0.00001,"SI","NO")')

        for i in range(2, n_iter + 1):
            r = i + 2
            k = i
            ws.cell(r, 1, k)
            ws.cell(r, 2, f"=IFERROR(F{r-1},B{r-1})")
            ws.cell(r, 3, f"=IFERROR({fx(f'B{r}')},0)")
            ws.cell(r, 4, f"=IFERROR(B{r-1},B{r-1})")
            ws.cell(r, 5, f"=IFERROR(C{r-1},0)")
            ws.cell(r, 6, f"=IFERROR(B{r}-(C{r}*(B{r}-D{r})/(C{r}-E{r})),B{r})")
            ws.cell(r, 7, f"=IFERROR(ABS((F{r}-B{r})/F{r})*100,0)")
            ws.cell(r, 8, f'=IF(IFERROR(G{r},0)<0.00001,"SI","NO")')

        footer_row = n_iter + 3
        _write_footer(
            ws, footer_row, f"A{footer_row}:H{footer_row}",
            f"MÉTODO DE LA SECANTE  —  {eq_label}", p,
            [[
                (1, "Ecuación:", True), (2, eq_label, False),
                (4, "x₀, x₁:", True), (5, f"{x0}, {x1}", False),
            ]]
        )
        _set_col_widths(ws, [5, 16, 16, 16, 16, 16, 12, 14])


# ---------------------------------------------------------------------------
# 14. VON MISES
# ---------------------------------------------------------------------------

class VonMisesTemplate:
    METHOD_KEY = "von_mises"
    HEADERS = ["k", "xₖ", "f(xₖ)", "f'(x₀) cte.", "xₖ₊₁", "Error %", "Convergencia"]
    N_COLS = 7

    def build(self, ws, fx_str: str, params: dict, n_iter: int = 25, eq_label: str = "",
              fpx_str: str = "") -> None:
        p = PALETTES[self.METHOD_KEY]
        x0 = params.get("x0", 1)

        _write_header(ws, self.HEADERS, p)
        _data_rows(ws, 2, n_iter + 1, p, self.N_COLS)

        fx = lambda ref: to_excel_formula(fx_str, ref)
        # f'(x₀) evaluated once and stored as constant
        fp_val = evaluate_derivative_at(fx_str, x0, order=1)
        if fp_val == int(fp_val):
            fp_val = int(fp_val)

        for i in range(n_iter):
            r = i + 2
            k = i
            ws.cell(r, 1, k)
            if k == 0:
                ws.cell(r, 2, x0)
            else:
                ws.cell(r, 2, f"=E{r-1}")

            bref = f"B{r}"
            ws.cell(r, 3, f"={fx(bref)}")
            ws.cell(r, 4, fp_val)   # constant derivative
            ws.cell(r, 5, f"={bref}-(C{r}/D{r})")
            ws.cell(r, 6, f"=ABS((E{r}-{bref})/E{r})*100")
            ws.cell(r, 7, f'=IF(F{r}<0.00001,"SI","NO")')

        footer_row = n_iter + 2
        _write_footer(
            ws, footer_row, f"A{footer_row}:G{footer_row}",
            f"MÉTODO DE VON MISES  —  {eq_label}", p,
            [[
                (1, "Ecuación:", True), (2, eq_label, False),
                (4, "f'(x₀):", True), (5, str(fp_val), False),
            ]]
        )
        _set_col_widths(ws, [5, 16, 16, 18, 16, 12, 14])


# ---------------------------------------------------------------------------
# Registry: map method key → template class
# ---------------------------------------------------------------------------

TEMPLATE_REGISTRY: dict[str, type] = {
    "biseccion":        BiseccionTemplate,
    "regula_falsi":     RegulaFalsiTemplate,
    "punto_fijo":       PuntoFijoTemplate,
    "aitken":           AitkenTemplate,
    "steffensen":       SteffensenTemplate,
    "newton_raphson":   NewtonRaphsonTemplate,
    "newton_modificado":NewtonModificadoTemplate,
    "newton_2do_orden": Newton2doOrdenTemplate,
    "chebyshev":        ChebyshevTemplate,
    "halley":           HalleyTemplate,
    "super_halley":     SuperHalleyTemplate,
    "ostrowsky":        OstrowskyTemplate,
    "secante":          SecanteTemplate,
    "von_mises":        VonMisesTemplate,
}